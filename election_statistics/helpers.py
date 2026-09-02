"""
Модуль вспомогательных функций и констант.

Утилиты для парсинга Excel, форматирования данных и настройки отчётов.
Не зависит от бизнес-логики и моделей: только константы и чистые функции.
"""

import re
from contextlib import contextmanager
from datetime import datetime
from typing import Any, Iterator, Optional

import openpyxl
from django.db.models import Q
from openpyxl.styles import Border, Side

# ==============================================================================
# Константы
# ==============================================================================

BAD_FORMAT = "Ошибка: документ не соответствует ожидаемому формату"

# Размер пачки для bulk_create / bulk_update в импортерах
BATCH = 500

# Подпись группы для цехов без производства в отчётах
NO_PRODUCTION = "Без производства"

# Соответствие заголовков колонок файла полям модели Employee
COLUMNS = {
    "Подразделение": "department",
    "Таб№": "tab_number",
    "Фамилия": "surname",
    "Имя": "name",
    "Отчество": "patronymic",
    "Должность": "position",
    "Категория": "category",
    "Дата рождения": "birth_date",
    "Регион": "region",
    "Город": "city",
    "Улица": "street",
    "Дом": "house",
    "УИК": "uik",
    "Адрес УИК": "uik_address",
    "Район": "district",
    "Округ": "okrug",
}

# Режимы отчёта по цеху: явка или выбор способа
REPORT_MODES = {
    "turnout": {
        "title": "Информация по голосованию на",
        "column": "Количество проголосовавших",
        "list_title": "Список НЕ принявших участие в голосовании:",
        "done": Q(voted=True),
    },
    "method": {
        "title": "Информация по выбору способа голосования на",
        "column": "Количество выбравших способ",
        "list_title": "Список НЕ выбравших способ голосования:",
        "done": ~Q(method=""),
    },
}

# Ширины колонок отчёта по цеху
REPORT_WIDTHS = (30, 40, 12)

# Символы, запрещённые в именах файлов архива
BAD_NAME_CHARS = re.compile(r"[\\/*?:\[\]]")

# ==============================================================================
# Утилиты парсинга и форматирования
# ==============================================================================


def _text(value: Any) -> str:
    """
    Приводит значение ячейки Excel к строке.

    Float без дробной части превращает в int-строку (Excel хранит
    табельные номера и УИКи как числа).

    Аргументы:
        value: значение ячейки.

    Возвращает:
        Строку без пробелов по краям; None даёт пустую строку.
    """
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _date(value: Any) -> Optional[datetime.date]:
    """
    Разбирает дату из ячейки Excel.

    Поддерживает datetime, date и строки "дд.мм.гггг" / "гггг-мм-дд".

    Аргументы:
        value: значение ячейки.

    Возвращает:
        date или None, если разобрать не удалось.
    """
    if isinstance(value, datetime):
        return value.date()
    if hasattr(value, "year"):
        return value
    head = _text(value).split()
    if not head:
        return None
    for fmt in ("%d.%m.%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(head[0], fmt).date()
        except ValueError:
            continue
    return None


@contextmanager
def _sheet(upload: Any) -> Iterator:
    """
    Контекст-менеджер для безопасного чтения Excel.

    Открывает книгу в read_only режиме и гарантированно закрывает
    ресурсы даже при ошибке внутри блока.

    Аргументы:
        upload: файл или байтовый поток с xlsx.

    Возвращает:
        Итератор строк листа (значениями).

    Исключения:
        ValueError: если файл не открывается как xlsx.
    """
    try:
        book = openpyxl.load_workbook(upload, read_only=True, data_only=True)
    except Exception as exc:
        # Пользователю — понятное сообщение, в трейсбек — настоящая причина
        raise ValueError(BAD_FORMAT) from exc

    rows = book.active.iter_rows(values_only=True)
    try:
        yield rows
    finally:
        rows.close()
        book.close()


def _header(rows: Iterator, wanted: dict) -> dict:
    """
    Ищет строку с заголовками, регистронезависимо.

    Строка считается заголовком, если совпала хотя бы одна колонка;
    наличие обязательных колонок проверяет вызывающий импортер.

    Аргументы:
        rows: итератор строк листа.
        wanted: словарь ожидаемых заголовков (например, COLUMNS).

    Возвращает:
        Словарь {имя_колонки: индекс}.

    Исключения:
        ValueError: если строка заголовков не найдена.
    """
    lookup = {name.casefold(): name for name in wanted}
    for row in rows:
        found = {}
        for index, cell in enumerate(row):
            key = _text(cell).casefold()
            if key in lookup:
                found[lookup[key]] = index
        if found:
            return found
    raise ValueError(BAD_FORMAT)


def _by_number(value: Any) -> tuple:
    """
    Ключ сортировки: числовые значения первыми по возрастанию,
    строки — после, в алфавитном порядке.

    Аргументы:
        value: номер цеха/УИКа.

    Возвращает:
        Кортеж для сортировки.
    """
    name = (value or "").strip()
    return (0, int(name), "") if name.isdigit() else (1, 0, name)


def padded_number(value: Any) -> str:
    """
    Дополняет числовой номер цеха нулями до 3 знаков.

    Аргументы:
        value: номер ("7" -> "007").

    Возвращает:
        Дополненный номер; нечисловые значения — без изменений.
    """
    name = (value or "").strip()
    return f"{int(name):03d}" if name.isdigit() else name


def department_file_name(department: str) -> str:
    """
    Формирует безопасное имя файла отчёта.

    Аргументы:
        department: номер цеха.

    Возвращает:
        Номер с нулями и заменой запрещённых символов на "-".
    """
    return BAD_NAME_CHARS.sub("-", padded_number(department))


def _apply_border(sheet: Any) -> None:
    """
    Применяет тонкую чёрную рамку ко всем заполненным ячейкам листа.

    Аргументы:
        sheet: лист openpyxl.
    """
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row_cells in sheet.iter_rows(
        min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column
    ):
        for cell in row_cells:
            cell.border = border


def _format_with_percent(count: int, total: int) -> str:
    """
    Форматирует число с процентом от общего количества.

    Аргументы:
        count: количество.
        total: всего (0 — процент не добавляется).

    Возвращает:
        Строку вида "5 (50%)".
    """
    if not total:
        return str(count)
    return f"{count} ({round(count / total * 100)}%)"
