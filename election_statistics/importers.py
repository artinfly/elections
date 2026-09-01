from typing import Any, Iterator, Tuple

from django.db.models import F
from django.utils import timezone

from .helpers import BATCH, COLUMNS, _date, _header, _sheet, _text
from .models import Employee


def _rows_by_tab(rows: Iterator, positions: dict) -> dict:
    """
    Разбирает строки файла в словарь {табельный номер: значения полей}.
    Пустые строки и строки без табельного номера пропускаются.
    """
    parsed = {}
    for row in rows:
        if not any(row):
            continue
        values = {}
        for name, index in positions.items():
            field = COLUMNS[name]
            cell = row[index] if index < len(row) else None
            values[field] = _date(cell) if field == "birth_date" else _text(cell)
        tab = values.pop("tab_number")
        if tab:
            parsed[tab] = values
    return parsed


def _known_rows(tabs: list, fields: list) -> dict:
    """
    Подтягивает из базы уже существующих сотрудников по табельным номерам.
    Запрос идёт чанками по 2000 номеров.
    """
    known = {}
    tabs = list(tabs)
    for start in range(0, len(tabs), 2000):
        chunk = tabs[start : start + 2000]
        for row in Employee.objects.filter(tab_number__in=chunk).values(
            "pk", "tab_number", *fields
        ):
            known[row.pop("tab_number")] = row
    return known


def import_base(upload: Any) -> Tuple[int, int, int]:
    """
    Импорт базы сотрудников из Excel-файла.
    Возвращает кортеж (создано, обновлено, всего строк).
    """
    with _sheet(upload) as rows:
        positions = _header(rows, COLUMNS)
        if "Таб№" not in positions:
            raise ValueError("Ошибка, документ не соответствует формату")
        parsed = _rows_by_tab(rows, positions)

    if not parsed:
        return 0, 0, 0

    fields = [COLUMNS[name] for name in positions if COLUMNS[name] != "tab_number"]
    known = _known_rows(parsed, fields)

    fresh, stale = [], []
    for tab, values in parsed.items():
        current = known.get(tab)
        if current is None:
            fresh.append(Employee(tab_number=tab, **values))
        elif any(current[field] != values[field] for field in fields):
            stale.append(Employee(pk=current["pk"], tab_number=tab, **values))

    if fresh:
        Employee.objects.bulk_create(fresh, batch_size=BATCH)
    if stale:
        Employee.objects.bulk_update(stale, fields, batch_size=BATCH)
    return len(fresh), len(stale), len(parsed)


def set_turnout(queryset, voted: bool = True):
    """
    Массово проставляет явку одним UPDATE-запросом.
    При voted=True ставит текущее время и копирует method в voted_method.
    """
    return queryset.update(
        voted=voted,
        voted_at=timezone.now() if voted else None,
        voted_method=F("method") if voted else "",
    )


def mark_voted(tabs: list, voted: bool = True) -> Tuple[int, int]:
    """
    Отмечает явку по списку табельных номеров.
    Возвращает (сколько отмечено, сколько не найдено).
    """
    tabs = {t for t in tabs if t}
    if not tabs:
        return 0, 0
    found = Employee.objects.filter(tab_number__in=tabs)
    missing = len(tabs) - found.count()
    return set_turnout(found, voted), missing


# ==============================================================================
# НОВЫЙ ИМПОРТ: Формат "Сводная таблица для отчета штабу"
# ==============================================================================

NEW_FORMAT_COLUMNS = {
    "Таб.№": "tab_number",
    "ФИО": "fio_raw",  # ФИО одной строкой, будем парсить отдельно
    "Очно на своем избирательном участке": "uik",
    "При помощи дистанционного электронного голосования ДЭГ": "deg",
    "Очно на временном избирательном участке на территории предприятия": "uvz",
    "Очно на избирательном участке 19 округа": "u19",
}


def _parse_fio(fio_str: str) -> Tuple[str, str, str]:
    """
    Разбирает ФИО из одной строки на surname, name, patronymic.
    Формат: "Фамилия Имя Отчество" или "Фамилия Имя"
    """
    parts = str(fio_str).strip().split()
    if len(parts) >= 3:
        return parts[0], parts[1], parts[2]
    elif len(parts) == 2:
        return parts[0], parts[1], ""
    elif len(parts) == 1:
        return parts[0], "", ""
    return "", "", ""


def _find_data_start(rows: Iterator) -> Tuple[Iterator, int]:
    """
    Ищет начало таблицы с данными (пропускает шапку с подразделением и ФИО ответственного).
    Возвращает итератор строк и номер строки с данными.
    """
    row_num = 0
    for row in rows:
        row_num += 1
        # Ищем строку с заголовками № п/п, Таб.№, ФИО
        cells = [str(cell or "").strip() for cell in row if cell is not None]
        if "№ п/п" in cells or "Таб.№" in cells or "ФИО" in cells:
            return rows, row_num
    return rows, 1


def import_voting_choices(upload: Any) -> Tuple[int, int, int]:
    """
    Импорт выборов способа голосования из нового формата Excel.
    Формат: "Сводная таблица для отчета штабу по голосованию"

    Структура файла:
    - Строка 1: Подразделение
    - Строка 2: ФИО ответственного
    - Строка 3-4: Заголовки
    - Строка 5+: Данные (№ п/п, Таб.№, ФИО, 4 колонки способов с цифрой "1")

    Логика:
    - Если в колонке стоит "1", значит выбран этот способ
    - ФИО разбивается на Фамилия, Имя, Отчество
    - Табельный номер - ключ поиска
    - Обновляется только поле method, остальные данные не трогаются

    Возвращает (обновлено, всего строк, ошибок)
    """
    from openpyxl import load_workbook

    try:
        book = load_workbook(upload, read_only=True, data_only=True)
        sheet = book.active
    except Exception:
        raise ValueError("Ошибка: не удалось открыть Excel файл")

    # Собираем все строки в память для многократного прохода
    all_rows = list(sheet.iter_rows(values_only=True))
    if not all_rows:
        raise ValueError("Ошибка: файл пустой")

    # Ищем заголовок (должен содержать "Таб.№" или "ФИО")
    header_row_idx = 0
    for idx, row in enumerate(all_rows):
        cells = [str(cell or "").strip() for cell in row if cell is not None]
        if "Таб.№" in cells or "ФИО" in cells:
            header_row_idx = idx
            break

    if header_row_idx >= len(all_rows) - 1:
        raise ValueError("Ошибка: не найдена строка с данными")

    # Находим индексы колонок
    header_row = all_rows[header_row_idx]
    col_indices = {}
    for idx, cell in enumerate(header_row):
        cell_str = str(cell or "").strip()
        for key in NEW_FORMAT_COLUMNS:
            if key in cell_str:
                col_indices[key] = idx
                break

    # Проверяем обязательные колонки
    if "Таб.№" not in col_indices:
        raise ValueError("Ошибка: не найдена колонка 'Таб.№'")
    if "ФИО" not in col_indices:
        raise ValueError("Ошибка: не найдена колонка 'ФИО'")

    # Проверяем колонки способов голосования
    method_cols = {
        "uik": col_indices.get("Очно на своем избирательном участке"),
        "deg": col_indices.get(
            "При помощи дистанционного электронного голосования ДЭГ"
        ),
        "uvz": col_indices.get(
            "Очно на временном избирательном участке на территории предприятия"
        ),
        "u19": col_indices.get("Очно на избирательном участке 19 округа"),
    }

    if not any(method_cols.values()):
        raise ValueError("Ошибка: не найдены колонки способов голосования")

    # Обрабатываем данные
    updated = 0
    total = 0
    errors = 0

    # Начинаем со строки после заголовка
    for row_idx in range(header_row_idx + 1, len(all_rows)):
        row = all_rows[row_idx]
        if not any(row):  # Пропускаем пустые строки
            continue

        total += 1

        try:
            # Получаем табельный номер
            tab_col = col_indices["Таб.№"]
            tab_number = _text(row[tab_col] if tab_col < len(row) else None)
            if not tab_number:
                continue  # Пропускаем строки без таб номера

            # Получаем ФИО и разбираем
            fio_col = col_indices["ФИО"]
            fio_raw = _text(row[fio_col] if fio_col < len(row) else None)
            surname, name, patronymic = _parse_fio(fio_raw)

            # Определяем способ голосования (ищем "1" в колонках)
            selected_method = ""
            for method_code, col_idx in method_cols.items():
                if col_idx is not None and col_idx < len(row):
                    cell_value = _text(row[col_idx])
                    if cell_value == "1":
                        if selected_method:  # Уже выбран один способ - ошибка
                            errors += 1
                            selected_method = ""
                            break
                        selected_method = method_code

            if not selected_method and not any(
                method_cols.get(m) is not None for m in method_cols
            ):
                continue  # Нет ни одного способа - пропускаем

            # Ищем сотрудника в базе и обновляем
            try:
                employee = Employee.objects.get(tab_number=tab_number)
                employee.method = selected_method
                employee.save(update_fields=["method"])
                updated += 1
            except Employee.DoesNotExist:
                # Сотрудник не найден - создаем минимальную запись
                employee = Employee(
                    tab_number=tab_number,
                    surname=surname,
                    name=name,
                    patronymic=patronymic,
                    method=selected_method,
                )
                employee.save()
                updated += 1

        except Exception as e:
            errors += 1
            continue

    book.close()
    return updated, total, errors
