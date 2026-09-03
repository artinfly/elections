"""
Модуль генерации стандартных Excel-отчётов и ZIP-архивов.

Описание:
    Содержит логику создания «бумажных» отчётов: разделение по производствам,
    способы голосования, сводные таблицы, полная выгрузка сотрудников,
    отчёты по конкретным цехам и упаковка этих отчётов в ZIP-архивы.
    Кастомный конструктор отчётов находится в файле custom_reports.py.
"""

from datetime import datetime
from typing import Any, Optional

import openpyxl
from django.db.models import Count, Q
from django.utils import timezone
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from utils.archiver import ReportArchiver

from .helpers import (
    COLUMNS,
    NO_PRODUCTION,
    REPORT_MODES,
    REPORT_WIDTHS,
    _apply_border,
    _by_number,
    _format_with_percent,
    department_file_name,
    padded_number,
)
from .models import DEG, METHOD_LABELS, UIK, UIK19, UVZ, Employee

# ==============================================================================
# Вспомогательные функции
# ==============================================================================


def _unique_name(taken: set, name: str) -> str:
    """
    Генерирует уникальное имя файла, избегая дубликатов в архиве.

    Описание:
        Если имя уже занято, добавляет суффикс "-2", "-3" и так далее,
        пока не будет найдено свободное имя. Регистрирует результат в наборе.

    Аргументы:
        taken: набор уже занятых имён (изменяется внутри функции).
        name: желаемое имя файла.

    Возвращает:
        str: гарантированно уникальное имя.
    """
    unique, attempt = name, 2
    while unique in taken:
        unique = f"{name}-{attempt}"
        attempt += 1
    taken.add(unique)
    return unique


def _share_row(
    sheet: Any, line: int, label: str, people: int, came: int, bold: Any = None
) -> int:
    """
    Записывает строку отчёта по производствам (всего / проголосовало / процент).

    Аргументы:
        sheet: лист книги openpyxl.
        line: номер строки для записи.
        label: подпись строки (название цеха, "Итого" или "Всего").
        people: общее количество сотрудников.
        came: количество проголосовавших.
        bold: объект жирного шрифта для итоговых строк (None для обычных).

    Возвращает:
        int: номер следующей свободной строки.
    """
    cells = (
        sheet.cell(line, 1, label),
        sheet.cell(line, 2, people),
        sheet.cell(line, 3, came),
        sheet.cell(line, 4, came / people if people else 0),
    )
    # Выравнивание числовых значений по центру.
    for cell in cells[1:]:
        cell.alignment = Alignment(horizontal="center")
    # Форматирование процента (например, 0.85 -> 85.00%).
    cells[3].number_format = "0.00%"

    if bold:
        for cell in cells:
            cell.font = bold
    return line + 1


def _method_share_row(
    sheet: Any,
    line: int,
    label: str,
    people: int,
    deg: int,
    uik: int,
    uvz: int,
    u19: int,
    bold: Any = None,
) -> int:
    """
    Записывает строку отчёта по способам голосования.

    Описание:
        Формирует строку с разбивкой по плановым способам голосования.
        Колонка "Итог" содержит сумму только тех, кто выбрал способ.

    Аргументы:
        sheet: лист книги openpyxl.
        line: номер строки для записи.
        label: подпись строки.
        people: общее количество сотрудников.
        deg: количество выбравших ДЭГ.
        uik: количество выбравших обычный УИК.
        uvz: количество выбравших УИК-УВЗ.
        u19: количество выбравших УИК-19.
        bold: объект жирного шрифта для итоговых строк.

    Возвращает:
        int: номер следующей свободной строки.
    """
    chosen = deg + uik + uvz + u19
    cells = (
        sheet.cell(line, 1, label),
        sheet.cell(line, 2, people),
        sheet.cell(line, 3, deg),
        sheet.cell(line, 4, uik),
        sheet.cell(line, 5, uvz),
        sheet.cell(line, 6, u19),
        sheet.cell(line, 7, chosen),
        sheet.cell(line, 8, chosen / people if people else 0),
    )
    for cell in cells[1:]:
        cell.alignment = Alignment(horizontal="center")
    cells[7].number_format = "0.00%"

    if bold:
        for cell in cells:
            cell.font = bold
    return line + 1


# ==============================================================================
# Отчёты по производствам
# ==============================================================================


def production_table() -> Any:
    """
    Отчёт "Разделение по производствам (Голосование)".

    Описание:
        Группирует цеха по производствам. Для каждого производства выводит
        список цехов с количеством сотрудников, проголосовавших и процентом явки.
        Внизу каждого производства строка "Итого", в самом конце — "Всего".

    Возвращает:
        Книга openpyxl с готовым отчётом.
    """
    # ВАЖНО: Бизнес-колонка «Производство» в отчётах строится по полю service.
    # Это историческое расхождение имён в структуре данных. Не переименовывать
    # на production без синхронного изменения логики импорта из внешней базы.
    grouped = {}
    for row in (
        Employee.objects.exclude(department="")
        .values("service", "department")
        .annotate(people=Count("id"), came=Count("id", filter=Q(voted=True)))
        .order_by()
    ):
        grouped.setdefault(row["service"] or NO_PRODUCTION, []).append(row)

    book = openpyxl.Workbook()
    sheet = book.active
    sheet.title = "По производствам"

    # Настройка ширины колонок.
    for index, width in enumerate((16, 14, 16, 10), 1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    # Фиксация заголовков при прокрутке.
    sheet.freeze_panes = "A3"

    bold = Font(bold=True)
    # Объединение ячеек для сложных заголовков.
    sheet.merge_cells("A1:A2")
    sheet.merge_cells("B1:B2")
    sheet.merge_cells("C1:D1")

    for coordinate, title in (
        ("A1", "Подразделение"),
        ("B1", "Общее число работающих"),
        ("C1", "Итог"),
        ("C2", "Количество проголосовавших"),
        ("D2", "%"),
    ):
        cell = sheet[coordinate]
        cell.value = title
        cell.font = bold
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )

    line = 3
    total_people = total_came = 0

    # Сортировка производств: обычные по алфавиту, "Без производства" всегда в конце.
    for production in sorted(grouped, key=lambda name: (name == NO_PRODUCTION, name)):
        # Заголовок производства (объединяем все колонки).
        sheet.merge_cells(start_row=line, start_column=1, end_row=line, end_column=4)
        sheet.cell(line, 1, production).font = bold
        line += 1

        people = came = 0
        # Сортировка цехов внутри производства: числовые номера первыми.
        for row in sorted(
            grouped[production], key=lambda item: _by_number(item["department"])
        ):
            line = _share_row(
                sheet,
                line,
                padded_number(row["department"]),
                row["people"],
                row["came"],
            )
            people += row["people"]
            came += row["came"]

        # Итог по производству.
        line = _share_row(sheet, line, "Итого", people, came, bold=bold)
        total_people += people
        total_came += came

    # Общий итог (с пустой строкой для визуального отделения).
    _share_row(sheet, line + 1, "Всего", total_people, total_came, bold=bold)
    return book


def production_method_table(exclude_u19: bool = False) -> Any:
    """
    Отчёт "Способы голосования по производствам".

    Описание:
        Аналогичен отчёту по явке, но вместо проголосовавших показывает
        распределение сотрудников по плановым способам голосования.

    Аргументы:
        exclude_u19: если True, сотрудники 19-го округа исключаются из отчёта.

    Возвращает:
        Книга openpyxl с готовым отчётом.
    """
    grouped = {}
    base_qs = Employee.objects.exclude(department="")
    if exclude_u19:
        base_qs = base_qs.exclude(okrug="19")

    for row in (
        base_qs.values("service", "department")
        .annotate(
            people=Count("id"),
            deg=Count("id", filter=Q(method=DEG)),
            uik=Count("id", filter=Q(method=UIK)),
            uvz=Count("id", filter=Q(method=UVZ)),
            u19=Count("id", filter=Q(method=UIK19)),
        )
        .order_by()
    ):
        grouped.setdefault(row["service"] or NO_PRODUCTION, []).append(row)

    book = openpyxl.Workbook()
    sheet = book.active
    # Ограничение имени листа до 31 символа (лимит Excel).
    sheet.title = (
        "Способы по производствам" if not exclude_u19 else "Способы (без 19)"
    )[:31]

    for index, width in enumerate((16, 10, 10, 12, 10, 16, 20, 12), 1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    sheet.freeze_panes = "A3"

    bold = Font(bold=True)
    sheet.merge_cells("A1:A2")
    sheet.merge_cells("B1:B2")
    sheet.merge_cells("C1:F1")
    sheet.merge_cells("G1:H1")

    for coordinate, title in (
        ("A1", "Подразделение"),
        ("B1", "Общее количество"),
        ("C1", "Способ голосования"),
        ("C2", "ДЭГ"),
        ("D2", "УИК"),
        ("E2", "УИК-УВЗ"),
        ("F2", "УИК-19"),
        ("G1", "Итог"),
        ("G2", "Кол-во зарегистрированных"),
        ("H2", "%"),
    ):
        cell = sheet[coordinate]
        cell.value = title
        cell.font = bold
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )

    line = 3
    totals = dict.fromkeys(("people", "deg", "uik", "uvz", "u19"), 0)

    for production in sorted(grouped, key=lambda name: (name == NO_PRODUCTION, name)):
        sheet.merge_cells(start_row=line, start_column=1, end_row=line, end_column=7)
        sheet.cell(line, 1, production).font = bold
        line += 1

        sub = dict.fromkeys(("people", "deg", "uik", "uvz", "u19"), 0)
        for row in sorted(
            grouped[production], key=lambda item: _by_number(item["department"])
        ):
            line = _method_share_row(
                sheet,
                line,
                padded_number(row["department"]),
                row["people"],
                row["deg"],
                row["uik"],
                row["uvz"],
                row["u19"],
            )
            for key in sub:
                sub[key] += row[key]

        line = _method_share_row(
            sheet,
            line,
            "Итого",
            sub["people"],
            sub["deg"],
            sub["uik"],
            sub["uvz"],
            sub["u19"],
            bold=bold,
        )
        for key in sub:
            totals[key] += sub[key]

    _method_share_row(
        sheet,
        line + 1,
        "Всего",
        totals["people"],
        totals["deg"],
        totals["uik"],
        totals["uvz"],
        totals["u19"],
        bold=bold,
    )
    return book


# ==============================================================================
# Сводная таблица
# ==============================================================================


def summary_table(
    group_field: str = "department",
    group_title: str = "Подразделение",
    exclude_u19: bool = False,
) -> Any:
    """
    Универсальная сводная таблица с группировкой по указанному полю.

    Описание:
        Формирует таблицу, где каждая строка — это группа (цех, УИК и т.д.).
        Показывает общее количество людей, план по способам голосования,
        фактическую явку и процент. Внизу строка "Итого".

    Аргументы:
        group_field: поле модели для группировки (например, "department" или "uik").
        group_title: заголовок первой колонки.
        exclude_u19: если True, сотрудники 19-го округа исключаются из отчёта.

    Возвращает:
        Книга openpyxl со сводной таблицей.
    """
    # Исключаем записи с пустым значением поля группировки.
    base_qs = Employee.objects.exclude(**{group_field: ""})
    if exclude_u19:
        base_qs = base_qs.exclude(okrug="19")

    # Сбрасываем дефолтную сортировку модели через .order_by(),
    # чтобы поля сортировки не попали в SQL-запрос GROUP BY.
    rows = sorted(
        base_qs.values(group_field)
        .annotate(
            people=Count("id"),
            plan_deg=Count("id", filter=Q(method=DEG)),
            plan_uik=Count("id", filter=Q(method=UIK)),
            plan_uvz=Count("id", filter=Q(method=UVZ)),
            plan_u19=Count("id", filter=Q(method=UIK19)),
            came=Count("id", filter=Q(voted=True)),
        )
        .order_by(),
        key=lambda row: _by_number(row[group_field]),
    )

    book = openpyxl.Workbook()
    sheet = book.active
    sheet.title = "Сводка"

    for index, width in enumerate((18, 16, 10, 10, 12, 16, 10), 1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    sheet.freeze_panes = "A2"

    bold = Font(bold=True)
    centered = Alignment(horizontal="center", wrap_text=True)
    headers = (
        group_title,
        "Количество людей",
        "ДЭГ",
        "УИК",
        "УИК-УВЗ",
        "УИК-19",
        "Проголосовавшие",
        "Процент",
    )
    for column, name in enumerate(headers, 1):
        cell = sheet.cell(1, column, name)
        cell.font = bold
        cell.alignment = centered

    totals = dict.fromkeys(
        ("people", "plan_deg", "plan_uik", "plan_uvz", "plan_u19", "came"), 0
    )
    line = 2
    for row in rows:
        values = (
            row["people"],
            row["plan_deg"],
            row["plan_uik"],
            row["plan_uvz"],
            row["plan_u19"],
            row["came"],
        )
        # Номера цехов дополняются до 3 цифр, прочие группы пишутся как есть.
        label = (
            padded_number(row[group_field])
            if group_field == "department"
            else row[group_field]
        )
        sheet.cell(line, 1, label)
        for shift, value in enumerate(values, 2):
            sheet.cell(line, shift, value).alignment = Alignment(horizontal="center")

        share = sheet.cell(line, 8, row["came"] / row["people"] if row["people"] else 0)
        share.number_format = "0.00%"
        share.alignment = Alignment(horizontal="center")

        for key in totals:
            totals[key] += row[key]
        line += 1

    # Итоговая строка.
    sheet.cell(line, 1, "Итого").font = bold
    for shift, key in enumerate(
        ("people", "plan_deg", "plan_uik", "plan_uvz", "plan_u19", "came"), 2
    ):
        cell = sheet.cell(line, shift, totals[key])
        cell.font = bold
        cell.alignment = Alignment(horizontal="center")

    share = sheet.cell(
        line, 8, totals["came"] / totals["people"] if totals["people"] else 0
    )
    share.font = bold
    share.number_format = "0.00%"
    share.alignment = Alignment(horizontal="center")

    return book


def summary_table_no_u19(
    group_field: str = "department", group_title: str = "Подразделение"
) -> Any:
    """
    Обёртка для обратной совместимости: сводная таблица без 19-го округа.

    Возвращает:
        Книга openpyxl со сводкой без сотрудников 19-го округа.
    """
    return summary_table(
        group_field=group_field, group_title=group_title, exclude_u19=True
    )


# ==============================================================================
# Полная выгрузка и отчёт по цеху
# ==============================================================================


def export_xlsx() -> Any:
    """
    Полная выгрузка всех сотрудников в один Excel-файл.

    Описание:
        Выгружает все поля из модели плюс информацию о плане, явке и месте голосования.
        Использует итератор для экономии оперативной памяти на больших базах.

    Возвращает:
        Книга openpyxl со всеми данными.
    """
    book = openpyxl.Workbook()
    sheet = book.active
    sheet.title = "Сотрудники"

    # Заголовки колонок.
    sheet.append(list(COLUMNS) + ["Способ (план)", "Проголосовал", "Где голосовал"])
    fields = list(COLUMNS.values())

    # Итератор с размером пачки 2000 строк не держит всю таблицу в памяти.
    for person in (
        Employee.objects.all()
        .order_by("department", "surname", "name", "patronymic")
        .iterator(chunk_size=2000)
    ):
        row = [getattr(person, f) for f in fields]
        row.append(METHOD_LABELS.get(person.method, ""))
        row.append("да" if person.voted else "нет")
        row.append(METHOD_LABELS.get(person.voted_method, ""))
        sheet.append(row)
    return book


def department_report(
    department: str, moment: Optional[datetime] = None, mode: str = "turnout"
) -> Any:
    """
    Формирует отдельный Excel-отчёт по одному цеху.

    Описание:
        Содержит шапку с датой, общую статистику (всего/сделано/процент)
        и список тех, у кого нужная отметка отсутствует (не проголосовали
        или не выбрали способ).

    Аргументы:
        department: номер цеха.
        moment: момент времени для шапки отчёта (по умолчанию текущее).
        mode: режим отчёта из REPORT_MODES — "turnout" (явка) или "method" (способ).

    Возвращает:
        Книга openpyxl с отчётом по цеху.
    """
    rule = REPORT_MODES[mode]
    moment = moment or timezone.localtime()
    people = Employee.objects.filter(department=department)
    total = people.count()
    done = people.filter(rule["done"]).count()

    book = openpyxl.Workbook()
    sheet = book.active
    sheet.title = f"Цех {department}"[:31]

    for index, width in enumerate(REPORT_WIDTHS, 1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    sheet.freeze_panes = "A6"

    bold = Font(bold=True)
    title = sheet.cell(1, 1, f"{rule['title']} {moment:%d.%m.%y %H:%M}")
    title.font = Font(bold=True, size=12)
    sheet.cell(2, 1, f"Цех {department}").font = bold

    for column, name in enumerate(
        ("Общее количество голосующих", rule["column"], "Процент"), 1
    ):
        cell = sheet.cell(4, column, name)
        cell.font = bold
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    sheet.cell(5, 1, total).alignment = Alignment(horizontal="center")
    sheet.cell(5, 2, done).alignment = Alignment(horizontal="center")
    share = sheet.cell(5, 3, (done / total) if total else 0)
    share.number_format = "0.00%"
    share.alignment = Alignment(horizontal="center")

    sheet.cell(7, 1, rule["list_title"]).font = bold

    # Список тех, у кого нужной отметки нет.
    row = 8
    for person in people.exclude(rule["done"]).order_by(
        "surname", "name", "patronymic"
    ):
        sheet.cell(row, 1, person.tab_number)
        sheet.cell(row, 2, person.fio)
        row += 1

    return book


# ==============================================================================
# Архивы по цехам
# ==============================================================================


def reports_archive(
    moment: Optional[datetime] = None, mode: str = "turnout"
) -> ReportArchiver:
    """
    Собирает ZIP-архив из отчётов по каждому цеху.

    Описание:
        Генерирует отдельный Excel-файл для каждого цеха и упаковывает их
        в один ZIP-архив. Используется для скачивания на странице экспорта.

    Аргументы:
        moment: момент времени для шапок отчётов.
        mode: режим отчёта — "turnout" (явка) или "method" (способ).

    Возвращает:
        ReportArchiver: наполненный сборщик архива (метод build() вызывается отдельно).
    """
    moment = moment or timezone.localtime()
    departments = (
        Employee.objects.exclude(department="")
        .values_list("department", flat=True)
        .distinct()
        .order_by()
    )
    archiver = ReportArchiver()
    taken = set()

    for department in sorted(departments, key=_by_number):
        unique = _unique_name(taken, department_file_name(department))
        archiver.add_workbook(
            department_report(department, moment, mode), f"{unique}.xlsx"
        )

    return archiver


def department_custom_report(
    department: str, params: dict, moment: Optional[datetime] = None
) -> Any:
    """
    Сводный отчёт по одному цеху (макет конструктора отчётов).

    Описание:
        Формирует таблицу с расширенной шапкой (группы колонок по способам голосования)
        для конкретного цеха. Применяет фильтры, выбранные пользователем в конструкторе отчетов.

    Аргументы:
        department: номер цеха.
        params: словарь фильтров конструктора (производство, способ, округ и т.д.).
        moment: момент времени для шапки отчёта.

    Возвращает:
        Книга openpyxl с отчётом по цеху.
    """
    # Локальный импорт: позволяет использовать функции конструктора отчетов,
    # избегая циклических зависимостей между модулями.
    from .custom_reports import _custom_qs, _draw_custom_groups

    moment = moment or timezone.localtime()

    # ИСПРАВЛЕНО: Применяем глобальные фильтры конструктора,
    # а затем ограничиваем выборку конкретным цехом.
    people = (
        _custom_qs(params)
        .filter(department=department)
        .order_by("surname", "name", "patronymic", "uik")
    )

    book = openpyxl.Workbook()
    sheet = book.active
    sheet.title = f"Цех {department}"[:31]
    bold = Font(bold=True)
    centered = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Ведущие колонки (шапка).
    column = 1
    for title in ("Номер УИК", "Цех", "Таб.№", "ФИО", "Округ"):
        sheet.merge_cells(
            start_row=1, start_column=column, end_row=2, end_column=column
        )
        cell = sheet.cell(1, column, title)
        cell.font = bold
        cell.alignment = centered
        column += 1

    # Отрисовка групп колонок (ДЭГ, УИК и т.д.) и получение списка предикатов.
    predicates = _draw_custom_groups(sheet, start_row=1, start_col=column)

    widths = (10, 8, 10, 42, 8) + (12,) * len(predicates)
    for index, width in enumerate(widths, 1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    sheet.freeze_panes = "A3"

    row = 3
    totals = [0] * len(predicates)
    total_people = 0

    # Итератор с размером пачки 2000 строк не держит всю выборку в памяти.
    for person in people.iterator(chunk_size=2000):
        total_people += 1
        sheet.cell(row, 1, person.uik)
        sheet.cell(row, 2, person.department)
        sheet.cell(row, 3, person.tab_number)
        sheet.cell(row, 4, person.fio)
        sheet.cell(row, 5, person.okrug)

        # Проставляем отметки "+" в соответствующие колонки.
        for offset, predicate in enumerate(predicates):
            if predicate(person):
                sheet.cell(row, 6 + offset, "+")
                totals[offset] += 1
        row += 1

    # Строка ИТОГО.
    sheet.cell(row, 1, "ИТОГО").font = bold
    sheet.cell(row, 2, total_people).font = bold
    sheet.cell(row, 2).alignment = centered

    for offset, total in enumerate(totals):
        cell = sheet.cell(row, 6 + offset, _format_with_percent(total, total_people))
        cell.font = bold
        cell.alignment = centered

    _apply_border(sheet)
    return book


def custom_reports_archive(
    params: dict, moment: Optional[datetime] = None
) -> ReportArchiver:
    """
    Собирает ZIP-архив сводных отчётов по каждому цеху (конструктор).

    Описание:
        Генерирует отчёты по цехам с применением фильтров конструктора
        и упаковывает их в ZIP-архив.

    Аргументы:
        params: фильтры конструктора (применяются к содержимому каждого файла).
        moment: момент времени для шапок отчётов.

    Возвращает:
        ReportArchiver: наполненный сборщик архива (метод build() вызывается отдельно).
    """
    moment = moment or timezone.localtime()
    departments = (
        Employee.objects.exclude(department="")
        .values_list("department", flat=True)
        .distinct()
        .order_by()
    )
    archiver = ReportArchiver()
    taken = set()

    for department in sorted(departments, key=_by_number):
        unique = _unique_name(taken, department_file_name(department))
        archiver.add_workbook(
            department_custom_report(department, params, moment), f"{unique}.xlsx"
        )

    return archiver
