import re
from collections import defaultdict
from contextlib import contextmanager
from datetime import datetime

import openpyxl
from django.db.models import Count, F, Q
from django.utils import timezone
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter

from utils.archiver import ReportArchiver

from .models import DEG, METHOD_LABELS, UIK, UIK19, UVZ, Employee

# Сообщение об ошибке, если загруженный файл не похож на ожидаемую таблицу
BAD_FORMAT = "Ошибка, документ не соответствует формату"

# Соответствие заголовков колонок Excel-файла полям модели Employee.
# Используется при импорте базы сотрудников.
COLUMNS = {
    "Подразделение": "department",
    "Таб№": "tab_number",
    "Фамилия": "surname",
    "Имя": "name",
    "Отчество": "patronymic",
    "Должность": "position",
    "Категория": "category",
    "Дата рождения": "birth_date",
    "Регион": "region",
    "Город": "city",
    "Улица": "street",
    "Дом": "house",
    "УИК": "uik",
    "Адрес УИК": "uik_address",
    "Район": "district",
    "Округ": "okrug",
}


def _text(value):
    """
    Приводит значение ячейки Excel к строке.
    Целочисленные float (например 123.0) превращает в "123" без хвоста ".0".
    """
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _date(value):
    """
    Разбирает дату рождения из ячейки Excel.
    Поддерживает datetime, date и строки в форматах "дд.мм.гггг" и "гггг-мм-дд".
    Возвращает date или None, если распознать не удалось.
    """
    if isinstance(value, datetime):
        return value.date()
    if hasattr(value, "year"):
        return value
    head = _text(value).split()
    if not head:
        return None
    for fmt in ("%d.%m.%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(head[0], fmt).date()
        except ValueError:
            continue
    return None


@contextmanager
def _sheet(upload):
    """
    Контекст-менеджер для чтения загруженного файла.
    Открывает книгу в режиме только чтения (read_only) и отдаёт итератор строк.
    Гарантированно закрывает книгу и строки даже при ошибке.
    """
    try:
        book = openpyxl.load_workbook(upload, read_only=True, data_only=True)
    except Exception:
        raise ValueError(BAD_FORMAT)
    rows = book.active.iter_rows(values_only=True)
    try:
        yield rows
    finally:
        rows.close()
        book.close()


def _header(rows, wanted):
    """
    Ищет строку с заголовками колонок в файле.
    Сравнивает ячейки без учёта регистра и возвращает словарь {заголовок: индекс колонки}.
    Если подходящая строка не найдена — файл не соответствует формату.
    """
    lookup = {name.casefold(): name for name in wanted}
    for row in rows:
        found = {}
        for index, cell in enumerate(row):
            key = _text(cell).casefold()
            if key in lookup:
                found[lookup[key]] = index
        if found:
            return found
    raise ValueError(BAD_FORMAT)


# Размер пачки для bulk_create / bulk_update при импорте
BATCH = 500


def _rows_by_tab(rows, positions):
    """
    Разбирает строки файла в словарь {табельный номер: значения полей}.
    Пустые строки и строки без табельного номера пропускаются.
    Если табномер встретился несколько раз, побеждает последняя строка.
    """
    parsed = {}
    for row in rows:
        if not any(row):
            continue
        values = {}
        for name, index in positions.items():
            field = COLUMNS[name]
            cell = row[index] if index < len(row) else None
            values[field] = _date(cell) if field == "birth_date" else _text(cell)
        tab = values.pop("tab_number")
        if tab:
            parsed[tab] = values
    return parsed


def _known_rows(tabs, fields):
    """
    Подтягивает из базы уже существующих сотрудников по табельным номерам.
    Запрос идёт чанками по 2000 номеров, чтобы не упереться в лимиты СУБД.
    Возвращает словарь {табельный номер: текущие значения полей из БД}.
    """
    known = {}
    tabs = list(tabs)
    for start in range(0, len(tabs), 2000):
        chunk = tabs[start : start + 2000]
        for row in Employee.objects.filter(tab_number__in=chunk).values(
            "pk", "tab_number", *fields
        ):
            known[row.pop("tab_number")] = row
    return known


def import_base(upload):
    """
    Импорт базы сотрудников из Excel-файла.
    Сравнивает строки файла с базой: новые записи создаёт (bulk_create),
    изменившиеся обновляет (bulk_update), неизменённые не трогает.
    Возвращает кортеж (создано, обновлено, всего строк).
    """
    with _sheet(upload) as rows:
        positions = _header(rows, COLUMNS)
        if "Таб№" not in positions:
            raise ValueError(BAD_FORMAT)
        parsed = _rows_by_tab(rows, positions)

    if not parsed:
        return 0, 0, 0

    fields = [COLUMNS[name] for name in positions if COLUMNS[name] != "tab_number"]
    known = _known_rows(parsed, fields)

    fresh, stale = [], []
    for tab, values in parsed.items():
        current = known.get(tab)
        if current is None:
            fresh.append(Employee(tab_number=tab, **values))
        elif any(current[field] != values[field] for field in fields):
            stale.append(Employee(pk=current["pk"], tab_number=tab, **values))

    if fresh:
        Employee.objects.bulk_create(fresh, batch_size=BATCH)
    if stale:
        Employee.objects.bulk_update(stale, fields, batch_size=BATCH)
    return len(fresh), len(stale), len(parsed)


def set_turnout(queryset, voted=True):
    """
    Массово проставляет явку одним UPDATE-запросом по всему queryset.
    При voted=True ставит текущее время и копирует запланированный способ в voted_method.
    При voted=False сбрасывает отметку (время и способ очищаются).
    """
    return queryset.update(
        voted=voted,
        voted_at=timezone.now() if voted else None,
        voted_method=F("method") if voted else "",
    )


def mark_voted(tabs, voted=True):
    """
    Отмечает явку (или снимает отметку) по списку табельных номеров.
    Возвращает кортеж (сколько отмечено, сколько номеров не найдено в базе).
    """
    tabs = {t for t in tabs if t}
    if not tabs:
        return 0, 0
    found = Employee.objects.filter(tab_number__in=tabs)
    missing = len(tabs) - found.count()
    return set_turnout(found, voted), missing


# Ширины колонок в отчёте по одному цеху
REPORT_WIDTHS = (14, 46, 10)

# Символы, недопустимые в именах файлов и листов Excel — заменяются на дефис
BAD_NAME_CHARS = re.compile(r"[\\/*?:\[\]]")

# Настройки двух режимов отчёта по цеху: явка (turnout) и выбор способа (method)
REPORT_MODES = {
    "turnout": {
        "title": "Информация по голосованию на",
        "column": "Количество проголосовавших",
        "list_title": "Список НЕ принявших участие в голосовании:",
        "done": Q(voted=True),
    },
    "method": {
        "title": "Информация по выбору способа голосования на",
        "column": "Количество выбравших способ",
        "list_title": "Список НЕ выбравших способ голосования:",
        "done": ~Q(method=""),
    },
}


def padded_number(value):
    """
    Дополняет числовой номер цеха нулями до трёх знаков ("7" -> "007").
    Нечисловые значения возвращает как есть.
    """
    name = (value or "").strip()
    return f"{int(name):03d}" if name.isdigit() else name


def department_file_name(department):
    """
    Формирует безопасное имя файла для отчёта по цеху (без запрещённых символов).
    """
    return BAD_NAME_CHARS.sub("-", padded_number(department))


def department_report(department, moment=None, mode="turnout"):
    """
    Формирует Excel-отчёт по одному цеху в одном из режимов (REPORT_MODES):
    шапка с датой, итоговые цифры (всего / сделано / процент) и список тех,
    кто ещё не проголосовал / не выбрал способ.
    """
    rule = REPORT_MODES[mode]
    moment = moment or timezone.localtime()
    people = Employee.objects.filter(department=department)
    total = people.count()
    done = people.filter(rule["done"]).count()

    book = openpyxl.Workbook()
    sheet = book.active
    sheet.title = BAD_NAME_CHARS.sub("-", f"Цех {department}")[:31]
    for index, width in enumerate(REPORT_WIDTHS, 1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    bold = Font(bold=True)
    title = sheet.cell(1, 1, f"{rule['title']} {moment:%d.%m.%y %H:%M}")
    title.font = Font(bold=True, size=12)
    sheet.cell(2, 1, f"Цех {department}").font = bold

    for column, name in enumerate(
        ("Общее количество голосующих", rule["column"], "Процент"), 1
    ):
        cell = sheet.cell(4, column, name)
        cell.font = bold
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    sheet.cell(5, 1, total).alignment = Alignment(horizontal="center")
    sheet.cell(5, 2, done).alignment = Alignment(horizontal="center")
    share = sheet.cell(5, 3, (done / total) if total else 0)
    share.number_format = "0.00%"
    share.alignment = Alignment(horizontal="center")

    sheet.cell(7, 1, rule["list_title"]).font = bold

    # Список тех, кто ещё не отметился
    row = 8
    for person in people.exclude(rule["done"]).order_by(
        "surname", "name", "patronymic"
    ):
        sheet.cell(row, 1, person.tab_number)
        sheet.cell(row, 2, person.fio)
        row += 1

    return book


def reports_archive(moment=None, mode="turnout"):
    """
    Собирает ZIP-архив из отчётов по каждому цеху (department_report).
    Следит за уникальностью имён файлов внутри архива (при совпадении добавляет -2, -3...).
    """
    moment = moment or timezone.localtime()
    departments = (
        Employee.objects.exclude(department="")
        .values_list("department", flat=True)
        .distinct()
        .order_by()
    )
    archiver = ReportArchiver()
    taken = set()
    for department in sorted(departments, key=_by_number):
        name = department_file_name(department)
        unique, attempt = name, 2
        while unique in taken:
            unique = f"{name}-{attempt}"
            attempt += 1
        taken.add(unique)
        archiver.add_workbook(
            department_report(department, moment, mode), f"{unique}.xlsx"
        )
    return archiver


# Ширины колонок сводной таблицы по цехам
SUMMARY_WIDTHS = (18, 16, 10, 10, 12, 12, 16, 10)


def _by_number(value):
    """
    Ключ сортировки: числовые номера цехов идут первыми и по возрастанию числа,
    остальные — следом в алфавитном порядке.
    """
    name = (value or "").strip()
    return (0, int(name), "") if name.isdigit() else (1, 0, name)


# Подпись для строк без привязки к производству
NO_PRODUCTION = "Без производства"
# Ширины колонок отчёта "Разделение по производствам"
PRODUCTION_WIDTHS = (16, 14, 16, 10)


def _share_row(sheet, line, label, people, came, bold=None):
    """
    Пишет строку отчёта по производствам: подразделение, всего, проголосовало, процент.
    Возвращает номер следующей строки.
    """
    cells = (
        sheet.cell(line, 1, label),
        sheet.cell(line, 2, people),
        sheet.cell(line, 3, came),
        sheet.cell(line, 4, came / people if people else 0),
    )
    for cell in cells[1:]:
        cell.alignment = Alignment(horizontal="center")
    cells[3].number_format = "0.00%"
    if bold:
        for cell in cells:
            cell.font = bold
    return line + 1


def production_table():
    """
    Отчёт "Разделение по производствам (Голосование)":
    цеха сгруппированы по производствам, для каждого — всего/проголосовало/процент,
    под каждым производством строка "Итого", в конце строка "Всего".
    """
    grouped = {}
    for row in (
        Employee.objects.exclude(department="")
        .values("service", "department")
        .annotate(people=Count("id"), came=Count("id", filter=Q(voted=True)))
        .order_by()
    ):
        grouped.setdefault(row["service"] or NO_PRODUCTION, []).append(row)

    book = openpyxl.Workbook()
    sheet = book.active
    sheet.title = "По производствам"
    for index, width in enumerate(PRODUCTION_WIDTHS, 1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    bold = Font(bold=True)
    sheet.merge_cells("A1:A2")
    sheet.merge_cells("B1:B2")
    sheet.merge_cells("C1:D1")
    for coordinate, title in (
        ("A1", "Подразделение"),
        ("B1", "Общее число работающих"),
        ("C1", "Итог"),
        ("C2", "Количество проголосовавших"),
        ("D2", "%"),
    ):
        cell = sheet[coordinate]
        cell.value = title
        cell.font = bold
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )

    line = 3
    total_people = total_came = 0
    for production in sorted(grouped, key=lambda name: (name == NO_PRODUCTION, name)):
        sheet.merge_cells(start_row=line, start_column=1, end_row=line, end_column=4)
        sheet.cell(line, 1, production).font = bold
        line += 1

        people = came = 0
        for row in sorted(
            grouped[production], key=lambda item: _by_number(item["department"])
        ):
            line = _share_row(
                sheet,
                line,
                padded_number(row["department"]),
                row["people"],
                row["came"],
            )
            people += row["people"]
            came += row["came"]

        line = _share_row(sheet, line, "Итого", people, came, bold=bold)
        total_people += people
        total_came += came

    _share_row(sheet, line + 1, "Всего", total_people, total_came, bold=bold)
    return book


# Ширины колонок отчёта "Способы голосования по производствам"
PRODUCTION_METHOD_WIDTHS = (16, 10, 10, 12, 10, 16, 20, 12)


def _method_share_row(sheet, line, label, people, deg, uik, uvz, u19, bold=None):
    """
    Пишет строку отчёта по способам голосования: подразделение, всего людей,
    counts по каждому способу, итого выбравших и процент.
    Возвращает номер следующей строки.
    """
    chosen = deg + uik + uvz + u19
    cells = (
        sheet.cell(line, 1, label),
        sheet.cell(line, 2, people),
        sheet.cell(line, 3, deg),
        sheet.cell(line, 4, uik),
        sheet.cell(line, 5, uvz),
        sheet.cell(line, 6, u19),
        sheet.cell(line, 7, chosen),
        sheet.cell(line, 8, chosen / people if people else 0),
    )
    for cell in cells[1:]:
        cell.alignment = Alignment(horizontal="center")
    cells[7].number_format = "0.00%"
    if bold:
        for cell in cells:
            cell.font = bold
    return line + 1


def production_method_table(exclude_u19=False):
    """
    Отчёт "Способы голосования по производствам": цеха сгруппированы по производствам,
    для каждого — количество людей и распределение по способам (ДЭГ, УИК, УВЗ, У19).
    Если exclude_u19=True, из выборки исключаются сотрудники 19 округа.
    """
    queryset = Employee.objects.exclude(department="")
    if exclude_u19:
        queryset = queryset.exclude(okrug="19")

    grouped = {}
    for row in (
        Employee.objects.exclude(department="")
        .values("service", "department")
        .annotate(
            people=Count("id"),
            deg=Count("id", filter=Q(method=DEG)),
            uik=Count("id", filter=Q(method=UIK)),
            uvz=Count("id", filter=Q(method=UVZ)),
            u19=Count("id", filter=Q(method=UIK19)),
        )
        .order_by()
    ):
        grouped.setdefault(row["service"] or NO_PRODUCTION, []).append(row)

    book = openpyxl.Workbook()
    sheet = book.active
    sheet.title = (
        "Способы по производствам" if not exclude_u19 else "Способы (без 19)"
    )[:31]
    for index, width in enumerate(PRODUCTION_METHOD_WIDTHS, 1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    bold = Font(bold=True)
    sheet.merge_cells("A1:A2")
    sheet.merge_cells("B1:B2")
    sheet.merge_cells("C1:F1")
    sheet.merge_cells("G1:H1")
    for coordinate, title in (
        ("A1", "Подразделение"),
        ("B1", "Общее количество"),
        ("C1", "Способ голосования"),
        ("C2", "ДЭГ"),
        ("D2", "УИК"),
        ("E2", "УИК-УВЗ"),
        ("F2", "УИК-19"),
        ("G1", "Итог"),
        ("G2", "Кол-во зарегестрированных"),
        ("H2", "%"),
    ):
        cell = sheet[coordinate]
        cell.value = title
        cell.font = bold
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )

    line = 3
    totals = dict.fromkeys(("people", "deg", "uik", "uvz", "u19"), 0)
    for production in sorted(grouped, key=lambda name: (name == NO_PRODUCTION, name)):
        sheet.merge_cells(start_row=line, start_column=1, end_row=line, end_column=7)
        sheet.cell(line, 1, production).font = bold
        line += 1

        sub = dict.fromkeys(("people", "deg", "uik", "uvz", "u19"), 0)
        for row in sorted(
            grouped[production], key=lambda item: _by_number(item["department"])
        ):
            line = _method_share_row(
                sheet,
                line,
                padded_number(row["department"]),
                row["people"],
                row["deg"],
                row["uik"],
                row["uvz"],
                row["u19"],
            )
            for key in sub:
                sub[key] += row[key]

        line = _method_share_row(
            sheet,
            line,
            "Итого",
            sub["people"],
            sub["deg"],
            sub["uik"],
            sub["uvz"],
            sub["u19"],
            bold=bold,
        )
        for key in sub:
            totals[key] += sub[key]

    _method_share_row(
        sheet,
        line + 1,
        "Всего",
        totals["people"],
        totals["deg"],
        totals["uik"],
        totals["uvz"],
        totals["u19"],
        bold=bold,
    )
    return book


def summary_table(group_field="department", group_title="Подразделение"):
    """
    Сводная таблица по цехам: строка на цех — всего людей, планы по способам,
    проголосовавшие и процент явки. Внизу строка "Итого".
    """
    rows = sorted(
        Employee.objects.exclude(**{group_field: ""})
        .values(group_field)
        .annotate(
            people=Count("id"),
            plan_deg=Count("id", filter=Q(method=DEG)),
            plan_uik=Count("id", filter=Q(method=UIK)),
            plan_uvz=Count("id", filter=Q(method=UVZ)),
            plan_u19=Count("id", filter=Q(method=UIK19)),
            came=Count("id", filter=Q(voted=True)),
        )
        .order_by(),
        key=lambda row: _by_number(row[group_field]),
    )

    book = openpyxl.Workbook()
    sheet = book.active
    sheet.title = "Сводка"
    for index, width in enumerate((18, 16, 10, 10, 12, 16, 10), 1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    bold = Font(bold=True)
    centered = Alignment(horizontal="center", wrap_text=True)
    headers = (
        group_title,
        "Количество людей",
        "ДЭГ",
        "УИК",
        "УИК-УВЗ",
        "УИК-19",
        "Проголосовавшие",
        "Процент",
    )
    for column, name in enumerate(headers, 1):
        cell = sheet.cell(1, column, name)
        cell.font = bold
        cell.alignment = centered

    totals = dict.fromkeys(
        ("people", "plan_deg", "plan_uik", "plan_uvz", "plan_u19", "came"), 0
    )
    line = 2
    for row in rows:
        values = (
            row["people"],
            row["plan_deg"],
            row["plan_uik"],
            row["plan_uvz"],
            row["plan_u19"],
            row["came"],
        )
        label = row[group_field]
        if group_field == "department":
            label = padded_number(label)
        sheet.cell(line, 1, label)
        for shift, value in enumerate(values, 2):
            sheet.cell(line, shift, value).alignment = Alignment(horizontal="center")
        share = sheet.cell(line, 8, row["came"] / row["people"] if row["people"] else 0)
        share.number_format = "0.00%"
        share.alignment = Alignment(horizontal="center")
        for key in totals:
            totals[key] += row[key]
        line += 1

    sheet.cell(line, 1, "Итого").font = bold
    for shift, key in enumerate(
        ("people", "plan_deg", "plan_uik", "plan_uvz", "plan_u19", "came"), 2
    ):
        cell = sheet.cell(line, shift, totals[key])
        cell.font = bold
        cell.alignment = Alignment(horizontal="center")
    share = sheet.cell(
        line, 8, totals["came"] / totals["people"] if totals["people"] else 0
    )
    share.font = bold
    share.number_format = "0.00%"
    share.alignment = Alignment(horizontal="center")

    return book


def summary_table_no_u19(group_field="department", group_title="Подразделение"):
    """
    Та же сводная таблица по цехам, но без сотрудников 19 округа.
    """
    rows = sorted(
        Employee.objects.exclude(**{group_field: ""})
        .exclude(okrug="19")
        .values(group_field)
        .annotate(
            people=Count("id"),
            plan_deg=Count("id", filter=Q(method=DEG)),
            plan_uik=Count("id", filter=Q(method=UIK)),
            plan_uvz=Count("id", filter=Q(method=UVZ)),
            plan_u19=Count("id", filter=Q(method=UIK19)),
            came=Count("id", filter=Q(voted=True)),
        )
        .order_by(),
        key=lambda row: _by_number(row[group_field]),
    )

    book = openpyxl.Workbook()
    sheet = book.active
    sheet.title = "Сводка"
    for index, width in enumerate(SUMMARY_WIDTHS, 1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    bold = Font(bold=True)
    centered = Alignment(horizontal="center", wrap_text=True)
    headers = (
        group_title,
        "Количество людей",
        "ДЭГ",
        "УИК",
        "УИК-УВЗ",
        "УИК-19",
        "Проголосовавшие",
        "Процент",
    )
    for column, name in enumerate(headers, 1):
        cell = sheet.cell(1, column, name)
        cell.font = bold
        cell.alignment = centered

    totals = dict.fromkeys(
        ("people", "plan_deg", "plan_uik", "plan_uvz", "plan_u19", "came"), 0
    )
    line = 2
    for row in rows:
        values = (
            row["people"],
            row["plan_deg"],
            row["plan_uik"],
            row["plan_uvz"],
            row["plan_u19"],
            row["came"],
        )
        label = row[group_field]
        if group_field == "department":
            label = padded_number(label)
        sheet.cell(line, 1, label)
        for shift, value in enumerate(values, 2):
            sheet.cell(line, shift, value).alignment = Alignment(horizontal="center")
        share = sheet.cell(line, 8, row["came"] / row["people"] if row["people"] else 0)
        share.number_format = "0.00%"
        share.alignment = Alignment(horizontal="center")
        for key in totals:
            totals[key] += row[key]
        line += 1

    sheet.cell(line, 1, "Итого").font = bold
    for shift, key in enumerate(
        ("people", "plan_deg", "plan_uik", "plan_uvz", "plan_u19", "came"), 2
    ):
        cell = sheet.cell(line, shift, totals[key])
        cell.font = bold
        cell.alignment = Alignment(horizontal="center")
    share = sheet.cell(
        line, 8, totals["came"] / totals["people"] if totals["people"] else 0
    )
    share.font = bold
    share.number_format = "0.00%"
    share.alignment = Alignment(horizontal="center")

    return book


def export_xlsx():
    """
    Полная выгрузка всех сотрудников в Excel: колонки из загрузки плюс
    способ голосования (план), отметка явки и фактическое место голосования.
    Читает базу итератором, чтобы не грузить всю таблицу в память.
    """
    book = openpyxl.Workbook()
    sheet = book.active
    sheet.title = "Сотрудники"
    sheet.append(list(COLUMNS) + ["Способ (план)", "Проголосовал", "Где голосовал"])
    fields = list(COLUMNS.values())
    for person in Employee.objects.all().iterator(chunk_size=2000):
        row = [getattr(person, f) for f in fields]
        row.append(METHOD_LABELS.get(person.method, ""))
        row.append("да" if person.voted else "нет")
        row.append(METHOD_LABELS.get(person.voted_method, ""))
        sheet.append(row)
    return book


# ---------------------------------------------------------------------------
# Сводный отчёт-конструктор (формируется по фильтрам из формы на странице export)
# ---------------------------------------------------------------------------

# Колонки сводного отчёта: (группа, подколонка, предикат от сотрудника).
# Группа "" — одиночная колонка без подколонок (например "Не пойдет").
# Вся семантика отчёта собрана здесь: если смысл колонки изменится,
# достаточно поправить предикат в этом списке.
# ВАЖНО: для "Не 19 округ" используется строгое вхождение в ["20", "21"],
# чтобы пустые округа не попадали в эту категорию.
CUSTOM_COLUMNS = [
    ("ДЭГ", "Планирует", lambda p: p.method == DEG),
    ("ДЭГ", "Проголосовал", lambda p: p.voted and p.voted_method == DEG),
    ("На участке", "Планирует", lambda p: p.method == UIK),
    ("На участке", "Проголосовал", lambda p: p.voted and p.voted_method == UIK),
    ("На участке УВЗ", "Планирует", lambda p: p.method == UVZ),
    ("На участке УВЗ", "Проголосовал", lambda p: p.voted and p.voted_method == UVZ),
    ("Не 19 округ", "Планирует", lambda p: p.okrug in ["20", "21"] and bool(p.method)),
    ("Не 19 округ", "Открепился", lambda p: p.okrug in ["20", "21"] and p.detached),
    ("Не 19 округ", "Проголосовал", lambda p: p.okrug in ["20", "21"] and p.voted),
    ("", "Не пойдет", lambda p: p.not_going),
]


def _custom_qs(params):
    """
    Строит QuerySet для сводного отчёта по параметрам формы-конструктора.
    Поддерживает фильтры: производство, служба, способ (план), цех, где голосовал,
    округ (включая спецзначения "none" — пустой и "20+21" — оба округа) и УИК.
    """
    qs = Employee.objects.all()
    production = (params.get("production") or "").strip()
    dep = (params.get("dep") or "").strip()
    method = (params.get("method") or "").strip()
    where = (params.get("where") or "").strip()
    okrug = (params.get("okrug") or "").strip()
    uik = (params.get("uik") or "").strip()
    service = (params.get("service") or "").strip()

    if production:
        qs = qs.filter(production=production)
    if service:
        qs = qs.filter(service=service)
    if dep:
        qs = qs.filter(department=dep)
    if method == "none":
        qs = qs.filter(method="")
    elif method:
        qs = qs.filter(method=method)
    if where == "none":
        qs = qs.filter(voted_method="")
    elif where:
        qs = qs.filter(voted_method=where)
    if okrug == "none":
        qs = qs.filter(okrug="")
    elif okrug == "20+21":
        qs = qs.filter(okrug__in=["20", "21"])
    elif okrug:
        qs = qs.filter(okrug=okrug)
    if uik:
        qs = qs.filter(uik=uik)
    return qs


def _custom_groups():
    """
    Группирует колонки CUSTOM_COLUMNS: [(имя группы, [(подколонка, предикат), ...]), ...].
    Порядок групп и колонок сохраняется как в CUSTOM_COLUMNS.
    """
    groups = []
    for group, sub, predicate in CUSTOM_COLUMNS:
        if groups and groups[-1][0] == group:
            groups[-1][1].append((sub, predicate))
        else:
            groups.append((group, [(sub, predicate)]))
    return groups


def _draw_custom_groups(sheet, start_row, start_col):
    """
    Рисует заголовки групп и подколонок из CUSTOM_COLUMNS начиная с указанной ячейки.
    Возвращает список предикатов в том же порядке, в каком они нарисованы.
    """
    bold = Font(bold=True)
    centered = Alignment(horizontal="center", vertical="center", wrap_text=True)

    column = start_col
    predicates = []
    for group, subs in _custom_groups():
        start, end = column, column + len(subs) - 1
        if group:
            sheet.merge_cells(
                start_row=start_row,
                start_column=start,
                end_row=start_row,
                end_column=end,
            )
            head = sheet.cell(start_row, column, group)
        else:
            sheet.merge_cells(
                start_row=start_row,
                start_column=start,
                end_row=start_row + 1,
                end_column=end,
            )
            head = sheet.cell(start_row, column, subs[0][0])
        head.font = bold
        head.alignment = centered
        for sub, predicate in subs:
            if group:
                cell = sheet.cell(start_row + 1, column, sub)
                cell.font = bold
                cell.alignment = centered
            predicates.append(predicate)
            column += 1
    return predicates


def _apply_border(sheet):
    """Применяет тонкую чёрную рамку ко всем ячейкам листа."""
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row_cells in sheet.iter_rows(
        min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column
    ):
        for cell in row_cells:
            cell.border = border


def custom_report(params, moment=None):
    """
    Формирует сводный Excel-отчёт по сотрудникам (режим "По людям") на основе фильтров формы.
    Каждый сотрудник занимает одну строку, в ячейках ставится "+" для выполненных предикатов.
    """
    moment = moment or timezone.localtime()
    people = _custom_qs(params).order_by(
        "uik", "department", "surname", "name", "patronymic"
    )

    book = openpyxl.Workbook()
    sheet = book.active
    sheet.title = "Сводный отчёт"
    bold = Font(bold=True)
    centered = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Ведущие колонки
    column = 1
    for title in ("Номер УИК", "Цех", "Таб.№", "ФИО", "Округ"):
        sheet.merge_cells(
            start_row=1, start_column=column, end_row=2, end_column=column
        )
        cell = sheet.cell(1, column, title)
        cell.font = bold
        cell.alignment = centered
        column += 1

    # Группы способов
    predicates = _draw_custom_groups(sheet, start_row=1, start_col=column)

    # Ширины колонок
    widths = (10, 8, 10, 42, 8) + (12,) * len(predicates)
    for index, width in enumerate(widths, 1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    row = 3
    totals = [0] * len(predicates)
    for person in people.iterator(chunk_size=2000):
        sheet.cell(row, 1, person.uik)
        sheet.cell(row, 2, person.department)
        sheet.cell(row, 3, person.tab_number)
        sheet.cell(row, 4, person.fio)
        sheet.cell(row, 5, person.okrug)
        for offset, predicate in enumerate(predicates):
            if predicate(person):
                sheet.cell(row, 6 + offset, "+")
                totals[offset] += 1
        row += 1

    # Строка ИТОГО
    sheet.cell(row, 1, "ИТОГО").font = bold
    for offset, total in enumerate(totals):
        cell = sheet.cell(row, 6 + offset, total)
        cell.font = bold
        cell.alignment = centered

    _apply_border(sheet)
    return book


def custom_production_summary(params, include_depts=False, moment=None):
    """
    Сводный отчёт с группировкой по производствам.
    Если include_depts=True, добавляется разбивка по цехам внутри каждого производства.
    Если include_depts=False, выводится одна строка на производство с общими итогами.
    Использует те же предикаты, что и custom_report.
    """
    moment = moment or timezone.localtime()
    people = _custom_qs(params).order_by(
        "production", "department", "surname", "name", "patronymic"
    )

    book = openpyxl.Workbook()
    sheet = book.active
    sheet.title = "Сводный отчёт"
    bold = Font(bold=True)
    centered = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Сначала собираем предикаты, чтобы знать общее число колонок
    predicates = []
    for group, subs in _custom_groups():
        for sub, predicate in subs:
            predicates.append(predicate)

    # Ведущие колонки
    if include_depts:
        lead_headers = ("Производство", "Цех", "Всего")
    else:
        lead_headers = ("Производство", "Всего")
    total_columns = len(lead_headers) + len(predicates)

    # Заголовок отчёта
    title = sheet.cell(
        1, 1, f"Итоговый отчёт по видам производства на {moment:%d.%m.%y %H:%M}"
    )
    title.font = Font(bold=True, size=12)
    title.alignment = centered
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_columns)

    # Заголовки ведущих колонок (строки 2-3)
    header_row = 2
    column = 1
    for title_text in lead_headers:
        sheet.merge_cells(
            start_row=header_row,
            start_column=column,
            end_row=header_row + 1,
            end_column=column,
        )
        cell = sheet.cell(header_row, column, title_text)
        cell.font = bold
        cell.alignment = centered
        column += 1

    # Группы способов (строки 2-3)
    _draw_custom_groups(sheet, start_row=header_row, start_col=column)

    # Ширины колонок
    if include_depts:
        widths = (16, 14, 10) + (12,) * len(predicates)
    else:
        widths = (24, 10) + (12,) * len(predicates)
    for index, width in enumerate(widths, 1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    # Группируем данные в памяти
    if include_depts:
        data = defaultdict(lambda: defaultdict(list))
    else:
        data = defaultdict(list)

    for person in people.iterator(chunk_size=2000):
        production = person.production or NO_PRODUCTION
        if include_depts:
            data[production][person.department].append(person)
        else:
            data[production].append(person)

    row = header_row + 2
    grand_total_people = 0
    grand_totals = [0] * len(predicates)

    for production in sorted(
        data.keys(), key=lambda name: (name == NO_PRODUCTION, name)
    ):
        if include_depts:
            # Заголовок производства
            sheet.merge_cells(
                start_row=row, start_column=1, end_row=row, end_column=total_columns
            )
            cell = sheet.cell(row, 1, production)
            cell.font = bold
            cell.alignment = Alignment(horizontal="center")
            row += 1

            prod_total_people = 0
            prod_totals = [0] * len(predicates)

            for department in sorted(data[production].keys(), key=_by_number):
                persons = data[production][department]
                count = len(persons)
                prod_total_people += count
                grand_total_people += count

                sheet.cell(row, 1, production)
                sheet.cell(row, 2, padded_number(department))
                sheet.cell(row, 3, count).alignment = Alignment(horizontal="center")

                for offset, predicate in enumerate(predicates):
                    val = sum(1 for p in persons if predicate(p))
                    if val > 0:
                        sheet.cell(row, 4 + offset, val).alignment = Alignment(
                            horizontal="center"
                        )
                    prod_totals[offset] += val
                    grand_totals[offset] += val
                row += 1

            # Итого по производству
            sheet.cell(row, 1, "")
            sheet.cell(row, 2, "Итого").font = bold
            sheet.cell(row, 3, prod_total_people).font = bold
            sheet.cell(row, 3).alignment = Alignment(horizontal="center")
            for offset, val in enumerate(prod_totals):
                cell = sheet.cell(row, 4 + offset, val)
                cell.font = bold
                cell.alignment = Alignment(horizontal="center")
            row += 1
        else:
            # Режим без цехов: одна строка на производство
            persons = data[production]
            count = len(persons)
            grand_total_people += count

            sheet.cell(row, 1, production)
            sheet.cell(row, 2, count).alignment = Alignment(horizontal="center")

            for offset, predicate in enumerate(predicates):
                val = sum(1 for p in persons if predicate(p))
                if val > 0:
                    sheet.cell(row, 3 + offset, val).alignment = Alignment(
                        horizontal="center"
                    )
                grand_totals[offset] += val
            row += 1

    # Всего по Обществу
    if include_depts:
        sheet.cell(row, 2, "Всего по Обществу").font = bold
        sheet.cell(row, 3, grand_total_people).font = bold
        sheet.cell(row, 3).alignment = Alignment(horizontal="center")
        for offset, val in enumerate(grand_totals):
            cell = sheet.cell(row, 4 + offset, val)
            cell.font = bold
            cell.alignment = Alignment(horizontal="center")
    else:
        sheet.cell(row, 1, "Всего по Обществу").font = bold
        sheet.cell(row, 2, grand_total_people).font = bold
        sheet.cell(row, 2).alignment = Alignment(horizontal="center")
        for offset, val in enumerate(grand_totals):
            cell = sheet.cell(row, 3 + offset, val)
            cell.font = bold
            cell.alignment = Alignment(horizontal="center")

    _apply_border(sheet)
    return book
