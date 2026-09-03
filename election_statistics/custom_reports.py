"""
Модуль генерации кастомных (динамических) отчётов по фильтрам конструктора.

Описание:
    Содержит логику создания сводных отчётов, настраиваемых пользователем
    через форму фильтров. Включает отчёт "по людям" (каждый сотрудник строкой)
    и отчёт "по производствам" (агрегированные данные).
    Вся бизнес-логика отображения галочек в колонках сосредоточена в списке CUSTOM_COLUMNS.
"""

from collections import defaultdict
from datetime import datetime
from typing import Any, Callable, Optional

import openpyxl
from django.db.models import Q
from django.utils import timezone
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from .helpers import (
    NO_PRODUCTION,
    _apply_border,
    _by_number,
    _format_with_percent,
    padded_number,
)
from .models import DEG, UIK, UIK19, UVZ, Employee

# ==============================================================================
# Конфигурация колонок сводного отчёта
# ==============================================================================

# Список колонок отчёта в формате: (Имя группы, Имя подколонки, Предикат-проверка).
# Предикат — это функция, которая принимает объект сотрудника (p) и возвращает True,
# если для него нужно поставить отметку в соответствующей колонке.
# Если бизнес-логика колонки изменится, править нужно только этот список.
CUSTOM_COLUMNS: list[tuple[str, str, Callable]] = [
    # Группа ДЭГ (дистанционное электронное голосование)
    ("ДЭГ", "Планирует", lambda p: p.method == DEG),
    ("ДЭГ", "Зарегистрирован", lambda p: p.mark_deg),
    ("ДЭГ", "Проголосовал", lambda p: p.voted and p.voted_method == DEG),
    # Группа голосования на обычном участке
    ("На участке", "Планирует", lambda p: p.method == UIK),
    ("На участке", "Проголосовал", lambda p: p.voted and p.voted_method == UIK),
    # Группа голосования на участке УВЗ (на предприятии)
    ("На участке УВЗ", "Планирует", lambda p: p.method == UVZ),
    ("На участке УВЗ", "Заявление оформил", lambda p: p.mark_uvz),
    ("На участке УВЗ", "Проголосовал", lambda p: p.voted and p.voted_method == UVZ),
    # Группа 19-го округа
    ("УИК-19", "Планирует", lambda p: p.method == UIK19),
    ("УИК-19", "Открепился", lambda p: p.method == UIK19 and p.detached),
    ("УИК-19", "Проголосовал", lambda p: p.voted and p.voted_method == UIK19),
    # Одиночные колонки без группы (имя группы пустое).
    # "Не определился" — сотрудник не выбрал способ.
    # "Отсутствовал по УП" — сотрудник отметил уважительную причину отсутствия.
    ("", "Не определился", lambda p: p.method not in (DEG, UIK, UVZ, UIK19)),
    ("", "Отсутствовал по УП", lambda p: p.absence),
]


# ==============================================================================
# Фильтрация и заголовки
# ==============================================================================


def _custom_qs(params: dict) -> Any:
    """
    Строит QuerySet сотрудников на основе параметров формы конструктора.

    Описание:
        Применяет фильтры, выбранные пользователем (производство, цех, способ,
        округ, УИК и т.д.), к общему списку сотрудников. Поддерживает специальные
        значения фильтров, такие как "Пусто" (none) и "20+21".

    Аргументы:
        params: словарь параметров из запроса (production, service, dep, method,
            where, okrug, uik).

    Возвращает:
        QuerySet: отфильтрованный список сотрудников.
    """
    qs = Employee.objects.all()

    production = (params.get("production") or "").strip()
    dep = (params.get("dep") or "").strip()
    method = (params.get("method") or "").strip()
    where = (params.get("where") or "").strip()
    okrug = (params.get("okrug") or "").strip()
    uik = (params.get("uik") or "").strip()
    service = (params.get("service") or "").strip()

    if production:
        qs = qs.filter(production=production)
    if service:
        qs = qs.filter(service=service)
    if dep:
        qs = qs.filter(department=dep)

    # Фильтр по запланированному способу голосования.
    if method == "none":
        qs = qs.filter(method="")
    elif method:
        qs = qs.filter(method=method)

    # Фильтр по фактическому месту голосования.
    if where == "none":
        qs = qs.filter(voted_method="")
    elif where:
        qs = qs.filter(voted_method=where)

    # Фильтр по избирательному округу.
    if okrug == "none":
        qs = qs.filter(okrug="")
    elif okrug == "20+21":
        qs = qs.filter(okrug__in=["20", "21"])
    elif okrug:
        qs = qs.filter(okrug=okrug)

    if uik:
        qs = qs.filter(uik=uik)

    return qs


def _custom_groups() -> list[tuple[str, list[tuple[str, Callable]]]]:
    """
    Группирует колонки из CUSTOM_COLUMNS по имени группы.

    Описание:
        Преобразует плоский список колонок в иерархическую структуру для отрисовки
        объединённых заголовков в Excel. Колонки с одинаковым именем группы
        (идущие подряд) объединяются в одну группу.

    ИСПРАВЛЕНО: Колонки с пустым именем группы (одиночные) больше не сливаются
    между собой, а остаются независимыми столбцами. Это позволяет корректно
    отображать несколько одиночных колонок подряд (например, "Не определился"
    и "Отсутствовал по УП").

    Возвращает:
        list: список кортежей (имя_группы, [(имя_подколонки, предикат), ...]).
    """
    groups = []
    for group, sub, predicate in CUSTOM_COLUMNS:
        # Пустая группа означает одиночную колонку без общего заголовка.
        # Такие колонки не должны объединяться с другими пустыми группами.
        if not group:
            groups.append((group, [(sub, predicate)]))
        # Если текущая колонка принадлежит той же группе, что и предыдущая,
        # добавляем её в конец списка подколонок этой группы.
        elif groups and groups[-1][0] == group:
            groups[-1][1].append((sub, predicate))
        else:
            # Иначе создаём новую группу.
            groups.append((group, [(sub, predicate)]))
    return groups


def _draw_custom_groups(sheet: Any, start_row: int, start_col: int) -> list[Callable]:
    """
    Отрисовывает объединённые заголовки групп и подколонки в листе Excel.

    Описание:
        Рисует шапку отчёта. Группы колонок объединяются по горизонтали в верхней строке.
        Подколонки рисуются во второй строке. Если имя группы пустое (одиночная колонка),
        ячейка объединяется по вертикали на две строки.

    Аргументы:
        sheet: объект листа openpyxl.
        start_row: номер строки, с которой начинается шапка.
        start_col: номер колонки, с которой начинается шапка.

    Возвращает:
        list: список предикатов в порядке их следования в колонках листа.
    """
    bold = Font(bold=True)
    centered = Alignment(horizontal="center", vertical="center", wrap_text=True)

    column = start_col
    predicates = []
    for group, subs in _custom_groups():
        start, end = column, column + len(subs) - 1
        if group:
            # Объединяем ячейки для имени группы по горизонтали.
            sheet.merge_cells(
                start_row=start_row,
                start_column=start,
                end_row=start_row,
                end_column=end,
            )
            head = sheet.cell(start_row, column, group)
        else:
            # Для одиночной колонки (пустая группа) объединяем по вертикали.
            sheet.merge_cells(
                start_row=start_row,
                start_column=start,
                end_row=start_row + 1,
                end_column=end,
            )
            head = sheet.cell(start_row, column, subs[0][0])

        head.font = bold
        head.alignment = centered

        # Отрисовка подколонок и сбор предикатов.
        for sub, predicate in subs:
            if group:
                cell = sheet.cell(start_row + 1, column, sub)
                cell.font = bold
                cell.alignment = centered
            predicates.append(predicate)
            column += 1
    return predicates


# ==============================================================================
# Отчёты
# ==============================================================================


def custom_report(params: dict) -> Any:
    """
    Генерирует сводный Excel-отчёт "По людям" на основе фильтров.

    Описание:
        Выводит каждого сотрудника отдельной строкой. В колонках проставляются
        отметки "+" в соответствии с предикатами из CUSTOM_COLUMNS.
        В конце добавляется строка ИТОГО с количеством и процентами.

    Аргументы:
        params: параметры фильтров конструктора.

    Возвращает:
        Книга openpyxl с готовым отчётом.
    """
    people = _custom_qs(params).order_by(
        "department", "surname", "name", "patronymic", "uik"
    )

    book = openpyxl.Workbook()
    sheet = book.active
    sheet.title = "Сводный отчёт"
    bold = Font(bold=True)
    centered = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Ведущие колонки (основная информация о сотруднике).
    column = 1
    for title in ("Номер УИК", "Цех", "Таб.№", "ФИО", "Округ"):
        sheet.merge_cells(
            start_row=1, start_column=column, end_row=2, end_column=column
        )
        cell = sheet.cell(1, column, title)
        cell.font = bold
        cell.alignment = centered
        column += 1

    # Отрисовка динамических колонок и получение списка предикатов.
    predicates = _draw_custom_groups(sheet, start_row=1, start_col=column)

    # Настройка ширины колонок.
    widths = (10, 8, 10, 42, 8) + (12,) * len(predicates)
    for index, width in enumerate(widths, 1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    # Фиксация заголовков при прокрутке.
    sheet.freeze_panes = "A3"

    row = 3
    totals = [0] * len(predicates)
    total_people = 0

    # Итератор с пачками по 2000 строк защищает от переполнения памяти.
    for person in people.iterator(chunk_size=2000):
        total_people += 1
        sheet.cell(row, 1, person.uik)
        sheet.cell(row, 2, person.department)
        sheet.cell(row, 3, person.tab_number)
        sheet.cell(row, 4, person.fio)
        sheet.cell(row, 5, person.okrug)

        # Простановка отметок в колонках.
        for offset, predicate in enumerate(predicates):
            if predicate(person):
                sheet.cell(row, 6 + offset, "+")
                totals[offset] += 1
        row += 1

    # Строка ИТОГО.
    sheet.cell(row, 1, "ИТОГО").font = bold
    sheet.cell(row, 2, total_people).font = bold
    sheet.cell(row, 2).alignment = centered

    for offset, total in enumerate(totals):
        cell = sheet.cell(row, 6 + offset, _format_with_percent(total, total_people))
        cell.font = bold
        cell.alignment = centered

    _apply_border(sheet)
    return book


def custom_production_summary(
    params: dict, include_depts: bool = False, moment: Optional[datetime] = None
) -> Any:
    """
    Генерирует сводный отчёт с группировкой по производствам.

    Описание:
        Агрегирует данные по производствам (опционально с разбивкой по цехам).
        Показывает количество сотрудников по каждому производству и процентные
        соотношения по колонкам из конструктора.

    Аргументы:
        params: параметры фильтров конструктора.
        include_depts: если True, выводит строки по цехам внутри каждого производства.
        moment: момент времени для заголовка отчёта.

    Возвращает:
        Книга openpyxl со сводным отчётом.
    """
    moment = moment or timezone.localtime()

    # Группировка и сортировка приведены к полю service.
    # Это обеспечивает единообразие со стандартными отчётами "по производствам"
    # в файле reports.py, которые исторически используют поле service.
    people = _custom_qs(params).order_by(
        "service", "department", "surname", "name", "patronymic"
    )

    book = openpyxl.Workbook()
    sheet = book.active
    sheet.title = "Сводный отчёт"
    bold = Font(bold=True)
    centered = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Собираем плоский список всех предикатов для отчёта.
    predicates = []
    for group, subs in _custom_groups():
        for sub, predicate in subs:
            predicates.append(predicate)

    # Заголовки ведущих колонок.
    lead_headers = (
        ("Производство", "Цех", "Всего") if include_depts else ("Производство", "Всего")
    )
    total_columns = len(lead_headers) + len(predicates)

    # Главный заголовок отчёта.
    title = sheet.cell(
        1, 1, f"Итоговый отчёт по видам производства на {moment:%d.%m.%y %H:%M}"
    )
    title.font = Font(bold=True, size=12)
    title.alignment = centered
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_columns)

    # Отрисовка заголовков колонок.
    header_row = 2
    column = 1
    for title_text in lead_headers:
        sheet.merge_cells(
            start_row=header_row,
            start_column=column,
            end_row=header_row + 1,
            end_column=column,
        )
        cell = sheet.cell(header_row, column, title_text)
        cell.font = bold
        cell.alignment = centered
        column += 1

    _draw_custom_groups(sheet, start_row=header_row, start_col=column)

    # Настройка ширины колонок.
    widths = (
        (16, 14, 10) + (12,) * len(predicates)
        if include_depts
        else (24, 10) + (12,) * len(predicates)
    )
    for index, width in enumerate(widths, 1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    sheet.freeze_panes = "A4"

    # Группировка данных в памяти для быстрого агрегирования.
    if include_depts:
        data = defaultdict(lambda: defaultdict(list))
    else:
        data = defaultdict(list)

    # Итератор с пачками по 2000 строк защищает от переполнения памяти.
    for person in people.iterator(chunk_size=2000):
        # Используем поле service для определения производства,
        # чтобы соответствовать стандартным отчётам.
        production_name = person.service or NO_PRODUCTION
        if include_depts:
            data[production_name][person.department].append(person)
        else:
            data[production_name].append(person)

    row = header_row + 2
    grand_total_people = 0
    grand_totals = [0] * len(predicates)

    # Сортировка производств: обычные по алфавиту, "Без производства" всегда в конце.
    for production in sorted(
        data.keys(), key=lambda name: (name == NO_PRODUCTION, name)
    ):
        if include_depts:
            # Заголовок производства (объединяем все колонки).
            sheet.merge_cells(
                start_row=row, start_column=1, end_row=row, end_column=total_columns
            )
            sheet.cell(row, 1, production).font = bold
            sheet.cell(row, 1).alignment = Alignment(horizontal="center")
            row += 1

            prod_total_people = 0
            prod_totals = [0] * len(predicates)

            # Сортировка цехов внутри производства.
            for department in sorted(data[production].keys(), key=_by_number):
                persons = data[production][department]
                count = len(persons)
                prod_total_people += count
                grand_total_people += count

                sheet.cell(row, 1, production)
                sheet.cell(row, 2, padded_number(department))
                sheet.cell(row, 3, count).alignment = Alignment(horizontal="center")

                # Подсчёт значений по предикатам для данного цеха.
                for offset, predicate in enumerate(predicates):
                    val = sum(1 for p in persons if predicate(p))
                    if val > 0:
                        sheet.cell(row, 4 + offset, val).alignment = Alignment(
                            horizontal="center"
                        )
                    prod_totals[offset] += val
                    grand_totals[offset] += val
                row += 1

            # Строка Итого по производству.
            sheet.cell(row, 2, "Итого").font = bold
            sheet.cell(row, 3, prod_total_people).font = bold
            sheet.cell(row, 3).alignment = Alignment(horizontal="center")
            for offset, val in enumerate(prod_totals):
                cell = sheet.cell(
                    row, 4 + offset, _format_with_percent(val, prod_total_people)
                )
                cell.font = bold
                cell.alignment = centered
            row += 1
        else:
            # Режим без разбивки по цехам.
            persons = data[production]
            count = len(persons)
            grand_total_people += count

            sheet.cell(row, 1, production)
            sheet.cell(row, 2, count).alignment = Alignment(horizontal="center")

            for offset, predicate in enumerate(predicates):
                val = sum(1 for p in persons if predicate(p))
                if val > 0:
                    sheet.cell(row, 3 + offset, val).alignment = Alignment(
                        horizontal="center"
                    )
                grand_totals[offset] += val
            row += 1

    # Строка Всего по Обществу.
    if include_depts:
        sheet.cell(row, 2, "Всего по Обществу").font = bold
        sheet.cell(row, 3, grand_total_people).font = bold
        sheet.cell(row, 3).alignment = Alignment(horizontal="center")
        for offset, val in enumerate(grand_totals):
            cell = sheet.cell(
                row, 4 + offset, _format_with_percent(val, grand_total_people)
            )
            cell.font = bold
            cell.alignment = centered
    else:
        sheet.cell(row, 1, "Всего по Обществу").font = bold
        sheet.cell(row, 2, grand_total_people).font = bold
        sheet.cell(row, 2).alignment = Alignment(horizontal="center")
        for offset, val in enumerate(grand_totals):
            cell = sheet.cell(
                row, 3 + offset, _format_with_percent(val, grand_total_people)
            )
            cell.font = bold
            cell.alignment = centered

    _apply_border(sheet)
    return book
