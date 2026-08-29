import io
import zipfile

import openpyxl
from django.contrib.auth.models import Group, User
from django.test import TestCase
from django.urls import get_resolver, reverse

from .models import DEG, UIK, UVZ, Employee
from .services import (
    COLUMNS,
    NO_PRODUCTION,
    custom_report,
    department_file_name,
    department_report,
    import_base,
    production_method_table,
    production_table,
    reports_archive,
    summary_table,
)


def _all_urls():
    """
    Возвращает список всех URL приложения, кроме login/logout
    (они доступны анонимам и проверяются отдельно).
    """
    return [
        reverse(pattern.name)
        for pattern in get_resolver("election_statistics.urls").url_patterns
        if pattern.name not in ("login", "logout")
    ]


# Эталонная строка сотрудника для тестов импорта
SAMPLE = {
    "department": "97",
    "tab_number": "0848103",
    "surname": "Иванов",
    "name": "Иван",
    "patronymic": "Иванович",
    "position": "Слесарь",
    "category": "Рабочий",
    "birth_date": "17.11.1975",
    "region": "Свердловская",
    "city": "Нижний Тагил",
    "street": "Ленина",
    "house": "1",
    "uik": "2632",
    "uik_address": "Школа 1",
    "district": "Дзержинский",
    "okrug": "21",
}


def _book(headers, rows):
    """Собирает тестовую xlsx-книгу в памяти из заголовков и строк."""
    book = openpyxl.Workbook()
    sheet = book.active
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    buffer = io.BytesIO()
    book.save(buffer)
    buffer.seek(0)
    return buffer


def _sample_file(*people):
    """Собирает тестовый файл в формате, который ожидает импорт (колонки COLUMNS)."""
    people = people or (SAMPLE,)
    return _book(
        list(COLUMNS),
        [[person.get(COLUMNS[name], "") for name in COLUMNS] for person in people],
    )


class AccessTests(TestCase):
    """Проверки доступа: анонимы, обычные пользователи и операторы."""

    def setUp(self):
        self.viewer = User.objects.create_user("viewer", password="x")
        self.operator = User.objects.create_user("operator", password="x")
        self.operator.groups.add(Group.objects.create(name="operator"))

    def test_anonymous_redirected_to_login(self):
        """Все страницы кроме login/logout редиректят анонима на страницу входа."""
        urls = _all_urls()
        # 20 маршрутов всего минус login и logout
        self.assertEqual(len(urls), 18)
        for url in urls:
            response = self.client.get(url)
            self.assertEqual(response.status_code, 302, url)
            self.assertIn("/login/", response["Location"], url)

    def test_logged_in_sees_pages(self):
        """Авторизованный пользователь открывает основные страницы."""
        self.client.force_login(self.viewer)
        for url in ["/", "/elections/", "/export/"]:
            self.assertEqual(self.client.get(url).status_code, 200, url)

    def test_upload_hidden_from_plain_user(self):
        """Обычному пользователю загрузка недоступна: страница и POST закрыты."""
        self.client.force_login(self.viewer)
        page = self.client.get("/upload/")
        self.assertContains(page, "Нет доступа")
        posted = self.client.post("/upload/base/", {})
        self.assertEqual(posted.status_code, 403)

    def test_operator_may_open_upload(self):
        """Оператор видит страницу загрузки и колонки файла."""
        self.client.force_login(self.operator)
        page = self.client.get("/upload/")
        self.assertContains(page, "Колонки файла")
        self.assertContains(page, "Таб№")

    def test_api_requires_login(self):
        """API без авторизации редиректит на логин."""
        self.assertEqual(
            self.client.post(
                "/api/method/", "{}", content_type="application/json"
            ).status_code,
            302,
        )


class ImportTests(TestCase):
    """Проверки импорта базы сотрудников из Excel."""

    def test_all_columns_are_parsed(self):
        """Все колонки файла корректно разбираются в поля модели."""
        created, updated, total = import_base(_sample_file())
        self.assertEqual((created, updated, total), (1, 0, 1))

        person = Employee.objects.get(tab_number="0848103")
        self.assertEqual(person.department, "97")
        self.assertEqual(person.surname, "Иванов")
        self.assertEqual(person.position, "Слесарь")
        self.assertEqual(str(person.birth_date), "1975-11-17")
        self.assertEqual(person.uik, "2632")
        self.assertEqual(person.okrug, "21")

    def test_second_import_keeps_method_and_voted(self):
        """Повторный импорт тем же файлом не затирает способ и явку."""
        import_base(_sample_file())
        Employee.objects.filter(tab_number="0848103").update(
            method=DEG, voted=True, voted_method=UVZ
        )

        created, updated, total = import_base(_sample_file())
        self.assertEqual((created, updated, total), (0, 0, 1))

        person = Employee.objects.get(tab_number="0848103")
        self.assertEqual(person.method, DEG)
        self.assertEqual(person.voted_method, UVZ)
        self.assertTrue(person.voted)

    def test_changed_row_is_counted_as_updated(self):
        """Изменившаяся строка файла считается обновлением, а не новым сотрудником."""
        import_base(_sample_file())
        moved = dict(SAMPLE, department="130")
        created, updated, total = import_base(_sample_file(moved))
        self.assertEqual((created, updated, total), (0, 1, 1))
        self.assertEqual(Employee.objects.get(tab_number="0848103").department, "130")

    def test_file_without_tab_number_is_refused(self):
        """Файл без колонки "Таб№" отклоняется."""
        with self.assertRaises(ValueError):
            import_base(_book(["Фамилия", "Имя"], [["Иванов", "Иван"]]))

    def test_broken_file_is_refused(self):
        """Битый файл отклоняется с ошибкой формата."""
        with self.assertRaises(ValueError):
            import_base(io.BytesIO(b"not a workbook"))


class VotedApiTests(TestCase):
    """Проверки API отметки явки (одиночной и массовой)."""

    def setUp(self):
        self.client.force_login(User.objects.create_user("u", password="x"))
        self.person = Employee.objects.create(
            tab_number="001", department="97", surname="Тест", name="Т"
        )

    def _post(self, payload):
        """Отправляет POST на /api/voted/ с JSON-телом."""
        return self.client.post("/api/voted/", payload, content_type="application/json")

    def test_place_is_taken_from_the_plan(self):
        """При отметке явки место голосования копируется из запланированного способа."""
        Employee.objects.filter(pk=self.person.pk).update(method=UVZ)
        self._post({"id": self.person.pk, "voted": True})
        self.person.refresh_from_db()
        self.assertTrue(self.person.voted)
        self.assertEqual(self.person.voted_method, UVZ)
        self.assertEqual(self.person.method, UVZ)

    def test_without_a_plan_the_turnout_is_refused(self):
        """Без запланированного способа явку проставить нельзя."""
        response = self._post({"id": self.person.pk, "voted": True})
        self.assertEqual(response.status_code, 400)
        self.person.refresh_from_db()
        self.assertFalse(self.person.voted)
        self.assertEqual(self.person.voted_method, "")

    def test_place_is_not_taken_from_the_request(self):
        """Место голосования нельзя подменить через запрос — только из плана."""
        Employee.objects.filter(pk=self.person.pk).update(method=UVZ)
        self._post({"id": self.person.pk, "voted": True, "method": DEG})
        self.person.refresh_from_db()
        self.assertEqual(self.person.voted_method, UVZ)

    def test_turnout_removal_clears_the_place(self):
        """Снятие отметки явки очищает место голосования, но план остаётся."""
        Employee.objects.filter(pk=self.person.pk).update(method=DEG)
        self._post({"id": self.person.pk, "voted": True})
        self._post({"id": self.person.pk, "voted": False})
        self.person.refresh_from_db()
        self.assertFalse(self.person.voted)
        self.assertEqual(self.person.voted_method, "")
        self.assertEqual(self.person.method, DEG)

    def test_bulk_marking_fills_places_from_plans(self):
        """Массовая отметка явки заполняет места из планов каждого сотрудника."""
        Employee.objects.filter(pk=self.person.pk).update(method=DEG)
        other = Employee.objects.create(
            tab_number="002", department="97", surname="Второй", name="В", method=UVZ
        )
        self.client.post(
            "/api/bulk-voted/",
            {"voted": True, "filters": {}},
            content_type="application/json",
        )
        self.person.refresh_from_db()
        other.refresh_from_db()
        self.assertEqual(self.person.voted_method, DEG)
        self.assertEqual(other.voted_method, UVZ)

    def test_bulk_removal_clears_the_place(self):
        """Массовое снятие явки очищает места голосования."""
        Employee.objects.filter(pk=self.person.pk).update(method=DEG)
        self._post({"id": self.person.pk, "voted": True})
        self.client.post(
            "/api/bulk-voted/",
            {"voted": False, "filters": {}},
            content_type="application/json",
        )
        self.person.refresh_from_db()
        self.assertFalse(self.person.voted)
        self.assertEqual(self.person.voted_method, "")

    def test_unknown_id_answers_404(self):
        """Несуществующий ID возвращает 404."""
        self.assertEqual(self._post({"id": 10**9, "voted": True}).status_code, 404)

    def test_unknown_plan_is_not_stored(self):
        """Неизвестный код способа не сохраняется в базу."""
        self.client.post(
            "/api/method/",
            {"id": self.person.pk, "method": "zzz"},
            content_type="application/json",
        )
        self.person.refresh_from_db()
        self.assertEqual(self.person.method, "")


class FilterTests(TestCase):
    """Проверки фильтров на страницах списков."""

    def setUp(self):
        self.client.force_login(User.objects.create_user("f", password="x"))
        Employee.objects.create(
            tab_number="p1", department="10", surname="Планер", name="И", method=DEG
        )
        Employee.objects.create(
            tab_number="p2",
            department="10",
            surname="Фактик",
            name="И",
            method=DEG,
            voted=True,
            voted_method=UVZ,
        )

    def test_plan_filter_works_on_both_pages(self):
        """Фильтр по запланированному способу работает на обеих страницах."""
        for url in ["/", "/elections/"]:
            page = self.client.get(url, {"method": DEG})
            self.assertEqual(page.context["found"], 2, url)
            page = self.client.get(url, {"method": UVZ})
            self.assertEqual(page.context["found"], 0, url)

    def test_place_filter(self):
        """Фильтр по фактическому месту голосования (включая "не указано")."""
        self.assertEqual(
            self.client.get("/elections/", {"where": UVZ}).context["found"], 1
        )
        self.assertEqual(
            self.client.get("/elections/", {"where": "none"}).context["found"], 1
        )


class ReportTests(TestCase):
    """Проверки отчёта по одному цеху и ZIP-архивов."""

    def setUp(self):
        self.client.force_login(User.objects.create_user("u", password="x"))
        for i in range(5):
            Employee.objects.create(
                tab_number=f"{i:06d}",
                department="001",
                surname=f"Первый{i}",
                name="И",
                voted=i < 3,
            )
        Employee.objects.create(
            tab_number="900001", department="097", surname="Девятый", name="П"
        )

    def test_archive_holds_a_file_per_department(self):
        """В архиве по файлу на каждый цех."""
        response = self.client.get("/export/archive/")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "application/zip")
        names = zipfile.ZipFile(io.BytesIO(response.content)).namelist()
        self.assertEqual(sorted(names), ["001.xlsx", "097.xlsx"])

    def test_report_matches_the_sample(self):
        """Отчёт по цеху совпадает с эталоном: шапка, цифры, список не проголосовавших."""
        book = department_report("001")
        sheet = book.active
        self.assertIn("Информация по голосованию", sheet.cell(1, 1).value)
        self.assertEqual(sheet.cell(2, 1).value, "Цех 001")
        self.assertEqual(sheet.cell(4, 1).value, "Общее количество голосующих")
        self.assertEqual(sheet.cell(5, 1).value, 5)
        self.assertEqual(sheet.cell(5, 2).value, 3)
        self.assertAlmostEqual(sheet.cell(5, 3).value, 0.6)
        self.assertEqual(sheet.cell(5, 3).number_format, "0.00%")
        self.assertIn("НЕ принявших участие", sheet.cell(7, 1).value)

        listed = [sheet.cell(r, 1).value for r in (8, 9)]
        self.assertEqual(listed, ["000003", "000004"])
        self.assertIsNone(sheet.cell(10, 1).value)

    def test_report_without_people_gives_zero_percent(self):
        """Пустой цех даёт процент 0, а не ошибку деления на ноль."""
        Employee.objects.filter(department="001").delete()
        self.assertEqual(department_report("001").active.cell(5, 3).value, 0)


class SummaryTests(TestCase):
    """Проверки сводной таблицы по цехам."""

    def setUp(self):
        self.client.force_login(User.objects.create_user("u", password="x"))
        Employee.objects.create(
            tab_number="a",
            department="7",
            surname="А",
            name="И",
            method=DEG,
            voted=True,
        )
        Employee.objects.create(
            tab_number="b", department="7", surname="Б", name="И", method=UVZ
        )
        Employee.objects.create(
            tab_number="c", department="130", surname="В", name="И", voted=True
        )
        Employee.objects.create(tab_number="d", department="97", surname="Г", name="И")

    def test_rows_sorted_as_numbers(self):
        """Числовые номера цехов сортируются как числа, с допадением до 3 знаков."""
        sheet = summary_table().active
        names = [sheet.cell(r, 1).value for r in range(2, sheet.max_row)]
        self.assertEqual(names, ["007", "097", "130"])

    def test_counts_and_total(self):
        """Счётчики по способам и строка "Итого" считаются верно."""
        sheet = summary_table().active
        first = [sheet.cell(2, c).value for c in range(1, 7)]
        self.assertEqual(first, ["007", 2, 1, 0, 1, 1])
        self.assertAlmostEqual(sheet.cell(2, 7).value, 0.5)
        self.assertEqual(sheet.cell(2, 7).number_format, "0.00%")

        last = sheet.max_row
        self.assertEqual(sheet.cell(last, 1).value, "Итого")
        self.assertEqual(
            [sheet.cell(last, c).value for c in range(2, 7)], [4, 1, 0, 1, 2]
        )
        self.assertAlmostEqual(sheet.cell(last, 7).value, 0.5)

    def test_grouping_field_is_a_parameter(self):
        """Группировку можно переключить (например, по УИК)."""
        sheet = summary_table(group_field="uik", group_title="УИК").active
        self.assertEqual(sheet.cell(1, 1).value, "УИК")

    def test_download(self):
        """Эндпоинт выгрузки сводки отвечает 200."""
        self.assertEqual(self.client.get("/export/summary/").status_code, 200)


class ArchiveModeTests(TestCase):
    """Проверки двух режимов архивов: явка и способы голосования."""

    def setUp(self):
        self.client.force_login(User.objects.create_user("u", password="x"))
        Employee.objects.create(
            tab_number="a",
            department="7",
            surname="А",
            name="И",
            method=DEG,
            voted=True,
        )
        Employee.objects.create(tab_number="b", department="7", surname="Б", name="И")

    def _sheet(self, url):
        """Скачивает архив и возвращает лист единственного файла внутри."""
        z = zipfile.ZipFile(io.BytesIO(self.client.get(url).content))
        self.assertEqual(z.namelist(), ["007.xlsx"])
        return openpyxl.load_workbook(io.BytesIO(z.read("007.xlsx"))).active

    def test_turnout_archive(self):
        """Архив по явке содержит отчёт о голосовании."""
        sheet = self._sheet("/export/archive/")
        self.assertIn("голосованию", sheet.cell(1, 1).value)
        self.assertEqual(sheet.cell(4, 2).value, "Количество проголосовавших")
        self.assertEqual(sheet.cell(5, 2).value, 1)

    def test_method_archive(self):
        """Архив по способам содержит отчёт о выборе способа."""
        sheet = self._sheet("/export/archive-methods/")
        self.assertIn("способа", sheet.cell(1, 1).value)
        self.assertEqual(sheet.cell(4, 2).value, "Количество выбравших способ")
        self.assertEqual(sheet.cell(5, 2).value, 1)

    def test_department_number_is_padded(self):
        """Номера цехов дополняются до 3 цифр, запрещённые символы заменяются."""
        self.assertEqual(department_file_name("7"), "007")
        self.assertEqual(department_file_name("130"), "130")
        self.assertEqual(department_file_name("цех/1"), "цех-1")


class UikStatsTests(TestCase):
    """Проверки API статистики по УИКам."""

    def setUp(self):
        self.client.force_login(User.objects.create_user("u", password="x"))
        plan = [
            ("2632", "10", True),
            ("2632", "10", False),
            ("2633", "20", True),
            ("", "20", False),
        ]
        for i, (uik, dep, voted) in enumerate(plan):
            Employee.objects.create(
                tab_number=f"s{i}",
                department=dep,
                surname=f"С{i}",
                name="И",
                uik=uik,
                voted=voted,
            )

    def test_breakdown_by_uik(self):
        """Разбивка по УИКам считает людей и явку."""
        rows = self.client.get("/api/uik-stats/").json()
        self.assertEqual(
            rows,
            [
                {"uik": "", "people": 1, "came": 0},
                {"uik": "2632", "people": 2, "came": 1},
                {"uik": "2633", "people": 1, "came": 1},
            ],
        )

    def test_breakdown_follows_the_filter(self):
        """Разбивка уважает переданные фильтры."""
        rows = self.client.get("/api/uik-stats/", {"dep": "10"}).json()
        self.assertEqual(rows, [{"uik": "2632", "people": 2, "came": 1}])


class CountsTests(TestCase):
    """Проверки счётчиков статистики на страницах и в API."""

    def setUp(self):
        self.user = User.objects.create_user("u", password="x")
        self.client.force_login(self.user)
        for i in range(4):
            Employee.objects.create(
                tab_number=f"t{i}",
                department="10" if i < 2 else "20",
                surname=f"Ф{i}",
                name="И",
                method=DEG if i == 0 else "",
            )

    def test_counts_follow_the_filter(self):
        """Счётчики на странице считаются по отфильтрованной выборке."""
        page = self.client.get("/", {"dep": "10"})
        counts = page.context["counts"]["method"]
        self.assertEqual(counts["total"], 2)
        self.assertEqual(counts["deg"], 1)

        page = self.client.get("/", {"dep": "20"})
        counts = page.context["counts"]["method"]
        self.assertEqual(counts["total"], 2)
        self.assertEqual(counts["deg"], 0)

    def test_api_returns_counts_for_the_same_filter(self):
        """API возвращает счётчики с теми же фильтрами, что пришли с фронта."""
        target = Employee.objects.get(tab_number="t2")
        response = self.client.post(
            "/api/method/",
            {"id": target.pk, "method": DEG, "filters": {"dep": "20"}},
            content_type="application/json",
        )
        body = response.json()
        self.assertEqual(body["method"]["total"], 2)
        self.assertEqual(body["method"]["deg"], 1)


class BadRequestTests(TestCase):
    """Проверки обработки некорректных запросов к API."""

    def setUp(self):
        self.client.force_login(User.objects.create_user("u", password="x"))

    def test_broken_json_answers_400(self):
        """Битый JSON возвращает 400 на всех изменяющих эндпоинтах."""
        for url in ["/api/method/", "/api/voted/", "/api/bulk-voted/"]:
            response = self.client.post(
                url, "{не json", content_type="application/json"
            )
            self.assertEqual(response.status_code, 400, url)

    def test_unknown_id_type_answers_404(self):
        """Не-числовой ID возвращает 404."""
        for url in ["/api/method/", "/api/voted/"]:
            response = self.client.post(
                url, {"id": "не число"}, content_type="application/json"
            )
            self.assertEqual(response.status_code, 404, url)


class PaginationTests(TestCase):
    """Проверки пагинации списков."""

    def setUp(self):
        self.client.force_login(User.objects.create_user("u", password="x"))
        for number in range(120):
            Employee.objects.create(
                tab_number=str(number), department="7", surname="А", name="И"
            )

    def test_page_is_not_repeated_in_the_link(self):
        """Параметр page не дублируется в ссылках пагинации."""
        page = self.client.get("/", {"dep": "7", "page": 2})
        self.assertEqual(page.context["query"], "dep=7")
        self.assertNotContains(page, "page=2&page=")


class ArchiveNameTests(TestCase):
    """Проверки уникальности имён файлов в архиве."""

    def test_departments_with_the_same_number_do_not_collide(self):
        """Цеха "7" и "007" дают разные имена файлов внутри архива."""
        for tab, department in (("a", "7"), ("b", "007")):
            Employee.objects.create(
                tab_number=tab, department=department, surname="А", name="И"
            )
        names = reports_archive().filenames
        self.assertEqual(len(names), 2)
        self.assertEqual(len(set(names)), 2)


class ProductionTests(TestCase):
    """Проверки отчёта "Разделение по производствам"."""

    def setUp(self):
        self.client.force_login(User.objects.create_user("u", password="x"))
        people = (
            ("a", "7", "Механосборочное", True),
            ("b", "7", "Механосборочное", False),
            ("c", "130", "Механосборочное", True),
            ("d", "44", "Инструментальное", False),
            ("e", "101", "", True),
        )
        for tab, department, production, voted in people:
            Employee.objects.create(
                tab_number=tab,
                department=department,
                production=production,
                surname="А",
                name="И",
                voted=voted,
            )

    def test_rows_are_grouped_by_production(self):
        """Цеха группируются по производствам, группы и итоги на своих местах."""
        sheet = production_table().active
        labels = [sheet.cell(row, 1).value for row in range(3, sheet.max_row + 1)]
        self.assertEqual(
            labels,
            [
                "Инструментальное",
                "044",
                "Итого",
                "Механосборочное",
                "007",
                "130",
                "Итого",
                NO_PRODUCTION,
                "101",
                "Итого",
                None,
                "Всего",
            ],
        )

    def test_totals_and_share(self):
        """Строка "Всего" считает людей, проголосовавших и процент."""
        sheet = production_table().active
        rows = {sheet.cell(row, 1).value: row for row in range(3, sheet.max_row + 1)}
        last = rows["Всего"]
        self.assertEqual(sheet.cell(last, 2).value, 5)
        self.assertEqual(sheet.cell(last, 3).value, 3)
        self.assertAlmostEqual(sheet.cell(last, 4).value, 0.6)
        self.assertEqual(sheet.cell(last, 4).number_format, "0.00%")

    def test_workshop_without_production_goes_to_the_last_group(self):
        """Цех без производства попадает в группу "Без производства" в конце."""
        sheet = production_table().active
        labels = [sheet.cell(row, 1).value for row in range(3, sheet.max_row + 1)]
        self.assertEqual(labels.index(NO_PRODUCTION), len(labels) - 5)

    def test_export_page_counts_only_filled_productions(self):
        """Счётчик производств на странице экспорта учитывает только заполненные."""
        page = self.client.get("/export/")
        self.assertEqual(page.context["productions_count"], 2)

    def test_download(self):
        """Эндпоинт выгрузки отчёта по производствам отвечает 200."""
        self.assertEqual(self.client.get("/export/productions/").status_code, 200)


class ProductionMethodsNoU19Tests(TestCase):
    """Проверки отчёта "Способы по производствам" с исключением 19 округа."""

    def setUp(self):
        self.client.force_login(User.objects.create_user("u", password="x"))
        # Сотрудник из 19 округа — в отчёт без u19 попасть не должен
        Employee.objects.create(
            tab_number="a",
            department="7",
            production="Механосборочное",
            surname="А",
            name="И",
            okrug="19",
            method=DEG,
        )
        # Сотрудник из 20 округа — попадает в оба варианта отчёта
        Employee.objects.create(
            tab_number="b",
            department="7",
            production="Механосборочное",
            surname="Б",
            name="И",
            okrug="20",
            method=UVZ,
        )

    def test_regular_report_counts_everyone(self):
        """Обычный отчёт учитывает оба округа."""
        sheet = production_method_table().active
        self.assertEqual(sheet.cell(sheet.max_row, 2).value, 2)

    def test_no_u19_report_skips_okrug_19(self):
        """Отчёт без 19 округа учитывает только сотрудников не из 19 округа."""
        sheet = production_method_table(exclude_u19=True).active
        last = sheet.max_row
        self.assertEqual(sheet.cell(last, 1).value, "Всего")
        self.assertEqual(sheet.cell(last, 2).value, 1)
        # Колонка УИК-УВЗ (5-я) — один человек с методом UVZ
        self.assertEqual(sheet.cell(last, 5).value, 1)

    def test_download(self):
        """Новый эндпоинт выгрузки отвечает 200."""
        self.assertEqual(
            self.client.get("/export/productions-methods-no-u19/").status_code, 200
        )


class CustomReportTests(TestCase):
    """Проверки сводного отчёта-конструктора (формируется по фильтрам)."""

    def setUp(self):
        self.client.force_login(User.objects.create_user("u", password="x"))
        # Округ 19, метод ДЭГ, проголосовал ДЭГ
        Employee.objects.create(
            tab_number="c1",
            department="10",
            production="П1",
            surname="А",
            name="И",
            okrug="19",
            uik="100",
            method=DEG,
            voted=True,
            voted_method=DEG,
        )
        # Округ 20, метод УВЗ, не голосует, помечен "не пойдет"
        Employee.objects.create(
            tab_number="c2",
            department="10",
            production="П1",
            surname="Б",
            name="И",
            okrug="20",
            uik="100",
            method=UVZ,
            not_going=True,
        )
        # Округ 21, метод УИК, без явки
        Employee.objects.create(
            tab_number="c3",
            department="20",
            production="П2",
            surname="В",
            name="И",
            okrug="21",
            uik="200",
            method=UIK,
        )

    def test_header_structure(self):
        """Шапка отчёта: ведущие колонки и группы способов на двух строках."""
        sheet = custom_report({}).active
        self.assertEqual(sheet.cell(1, 1).value, "Номер УИК")
        self.assertEqual(sheet.cell(1, 6).value, "ДЭГ")
        self.assertEqual(sheet.cell(2, 6).value, "Планирует")
        self.assertEqual(sheet.cell(2, 7).value, "Проголосовал")
        # Одиночная колонка без группы
        self.assertEqual(sheet.cell(1, 15).value, "Не пойдет")

    def test_marks_and_totals(self):
        """Отметки-единицы расставляются по предикатам, ИТОГО считает суммы."""
        sheet = custom_report({}).active
        # 3 сотрудника + строка ИТОГО, данные с 3-й строки
        self.assertEqual(sheet.max_row, 3 + 3 + 1)
        last = sheet.max_row
        self.assertEqual(sheet.cell(last, 1).value, "ИТОГО")
        # Суммы по колонкам: ДЭГ план/факт, УИК план/факт, УВЗ план/факт,
        # не-19 план/открепился/факт, не пойдет
        totals = [sheet.cell(last, c).value for c in range(6, 16)]
        self.assertEqual(totals, [1, 1, 1, 0, 1, 0, 2, 0, 0, 1])

    def test_okrug_20_plus_21_filter(self):
        """Фильтр "20+21" берёт оба округа."""
        sheet = custom_report({"okrug": "20+21"}).active
        self.assertEqual(sheet.max_row, 3 + 2 + 1)

    def test_okrug_none_filter(self):
        """Фильтр "Пусто" не находит никого (у всех округ заполнен)."""
        sheet = custom_report({"okrug": "none"}).active
        self.assertEqual(sheet.max_row, 3)

    def test_production_filter(self):
        """Фильтр по производству сужает выборку."""
        sheet = custom_report({"production": "П2"}).active
        self.assertEqual(sheet.max_row, 3 + 1 + 1)

    def test_download(self):
        """Эндпоинт выгрузки сводного отчёта отвечает 200."""
        self.assertEqual(self.client.get("/export/custom/").status_code, 200)
